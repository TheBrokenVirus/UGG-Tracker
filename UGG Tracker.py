"""
Roblox XP Tracker Discord Bot
------------------------------
Since the bot can't read XP directly from the game, users self-report their
current total XP with /checkin. The bot stores each checkin with a timestamp
and derives: gain since last checkin, average daily rate, projected weekly
total, and whether the user is on pace to hit the weekly requirement
(default 20,000 XP/week).

Data is stored in a local JSON file (xp_data.json), scoped per-guild.
"""

import os
import csv
import copy
import io
import json
from pathlib import Path
from typing import Optional
from datetime import datetime, timedelta, timezone

import discord
from discord import app_commands
from discord.ext import commands, tasks

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

try:
    import openpyxl  # optional — only needed for .xlsx import
except ImportError:
    openpyxl = None

try:
    import matplotlib
    matplotlib.use("Agg")  # headless backend — no display available on a server
    import matplotlib.pyplot as plt
    import matplotlib.dates as mdates
except ImportError:
    plt = None

DATA_FILE = Path(__file__).parent / "xp_data.json"
DEFAULT_WEEKLY_REQUIREMENT = 20000
MAX_CHECKINS_STORED = 50  # per user, to keep the file from growing forever

# ---------------------------------------------------------------------------
# Time helpers
# ---------------------------------------------------------------------------

def utcnow() -> datetime:
    return datetime.now(timezone.utc)

def iso(dt: datetime) -> str:
    return dt.isoformat()

def parse_iso(s: str) -> datetime:
    return datetime.fromisoformat(s)

def format_duration(hours: float) -> str:
    """Formats a fractional hour count into a 'Xd Yh Zm' style string,
    dropping any zero-valued components (e.g. '3d 12h 30m', '2h 5m', '45m').
    Works for durations of any length — minutes, hours, or multiple days."""
    total_minutes = round(hours * 60)
    d, rem = divmod(total_minutes, 1440)
    h, m = divmod(rem, 60)
    parts = []
    if d:
        parts.append(f"{d}d")
    if h:
        parts.append(f"{h}h")
    if m or not parts:
        parts.append(f"{m}m")
    return " ".join(parts)

def format_timedelta(delta: timedelta) -> str:
    """Same as format_duration, but takes a timedelta directly."""
    return format_duration(delta.total_seconds() / 3600)

def week_range_str(week_start_dt: datetime) -> str:
    """Formats a week's date range using Discord timestamp tags, which
    render in each viewer's own local timezone automatically."""
    week_end_dt = week_start_dt + timedelta(days=7)
    start_ts, end_ts = int(week_start_dt.timestamp()), int(week_end_dt.timestamp())
    return f"<t:{start_ts}:D> → <t:{end_ts}:D>"

def build_progress_bar(current: float, target: float, length: int = 14) -> str:
    """Renders a simple text progress bar using block characters, e.g.
    '████████░░░░░░ 57%'. Clamps at 100% even if current exceeds target
    (finishing early shouldn't render a broken/overflowing bar)."""
    if target <= 0:
        pct = 1.0 if current > 0 else 0.0
    else:
        pct = max(min(current / target, 1.0), 0.0)
    filled = round(pct * length)
    bar = "█" * filled + "░" * (length - filled)
    return f"{bar} {pct * 100:.0f}%"

# ---------------------------------------------------------------------------
# Storage
# ---------------------------------------------------------------------------

def load_data() -> dict:
    if DATA_FILE.exists():
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return {"guilds": {}}

def save_data(data: dict) -> None:
    with open(DATA_FILE, "w") as f:
        json.dump(data, f, indent=2)

DEFAULT_PRORATE_THRESHOLD_HOURS = 168.0  # 7 days = always prorate any lateness (original behavior)
DEFAULT_INACTIVITY_THRESHOLD_HOURS = 24.0  # ping people who haven't checked in when this much time is left

def default_guild_data() -> dict:
    """The default shape for a fresh guild entry. Also used by /restore to
    fill in any keys missing from an older or hand-edited backup file."""
    return {
        "requirement": DEFAULT_WEEKLY_REQUIREMENT,
        "users": {},
        "week_anchor": None,
        "show_recent_rate": True,
        "compact_leaderboard": False,
        "prorate_threshold_hours": DEFAULT_PRORATE_THRESHOLD_HOURS,
        "xp_roles": [],
        "restricted_channel_id": None,
        "announce_channel_id": None,
        "weekly_post_channel_id": None,
        "last_weekly_post_marker": None,
        "inactivity_channel_id": None,
        "inactivity_threshold_hours": DEFAULT_INACTIVITY_THRESHOLD_HOURS,
        "last_inactivity_ping_marker": None,
        "inactivity_include_behind_pace": False,
        "leaderboard_pagination": False,
        "admin_log_channel_id": None,
    }

def get_guild(data: dict, guild_id: int) -> dict:
    gid = str(guild_id)
    if gid not in data["guilds"]:
        data["guilds"][gid] = default_guild_data()
    return data["guilds"][gid]

def normalize_guild_data(guild_data: dict) -> dict:
    """Fills in any keys missing from a guild_data dict (e.g. one restored
    from an older backup taken before some settings existed) using current
    defaults, without touching keys that are already present."""
    merged = default_guild_data()
    merged.update(guild_data)
    return merged

def get_restricted_channel_id(guild_data: dict) -> Optional[int]:
    """Backward-compatible getter — older guild entries won't have this key."""
    return guild_data.get("restricted_channel_id")

def get_admin_log_channel_id(guild_data: dict) -> Optional[int]:
    """Backward-compatible getter — older guild entries won't have this key."""
    return guild_data.get("admin_log_channel_id")

def inactivity_includes_behind_pace(guild_data: dict) -> bool:
    """Backward-compatible getter — older guild entries won't have this key."""
    return guild_data.get("inactivity_include_behind_pace", False)

def get_announce_channel_id(guild_data: dict) -> Optional[int]:
    """Backward-compatible getter — older guild entries won't have this key."""
    return guild_data.get("announce_channel_id")

def get_weekly_post_channel_id(guild_data: dict) -> Optional[int]:
    return guild_data.get("weekly_post_channel_id")

def get_inactivity_channel_id(guild_data: dict) -> Optional[int]:
    return guild_data.get("inactivity_channel_id")

def get_inactivity_threshold_hours(guild_data: dict) -> float:
    return guild_data.get("inactivity_threshold_hours", DEFAULT_INACTIVITY_THRESHOLD_HOURS)

def leaderboard_pagination_enabled(guild_data: dict) -> bool:
    return guild_data.get("leaderboard_pagination", False)

def require_tracking_channel():
    """Command decorator: if the server has designated a channel with
    /setchannel, restricts the command to that channel only. If no
    channel is set, every channel is allowed (no behavior change)."""
    async def predicate(interaction: discord.Interaction) -> bool:
        data = load_data()
        guild_data = get_guild(data, interaction.guild_id)
        channel_id = get_restricted_channel_id(guild_data)
        if channel_id is None or interaction.channel_id == channel_id:
            return True
        raise app_commands.CheckFailure(f"This command can only be used in <#{channel_id}>.")
    return app_commands.check(predicate)

def show_recent_rate_enabled(guild_data: dict) -> bool:
    """Backward-compatible getter — older guild entries won't have this key."""
    return guild_data.get("show_recent_rate", True)

def compact_leaderboard_enabled(guild_data: dict) -> bool:
    """Backward-compatible getter — older guild entries won't have this key."""
    return guild_data.get("compact_leaderboard", False)

def get_prorate_threshold_hours(guild_data: dict) -> float:
    """Backward-compatible getter — older guild entries won't have this key.
    This is how many hours of the week must remain (or less) at someone's
    first checkin for proration to kick in. Default is the full week, i.e.
    any lateness at all triggers it."""
    return guild_data.get("prorate_threshold_hours", DEFAULT_PRORATE_THRESHOLD_HOURS)

def get_xp_roles(guild_data: dict) -> list:
    """Backward-compatible getter — older guild entries won't have this key.
    Returns a list of {"xp": int, "role_id": int, "exclusive": bool}
    mappings, sorted ascending. "exclusive" defaults to True for entries
    saved before this flag existed, preserving old swap-out behavior."""
    roles = guild_data.get("xp_roles", [])
    for r in roles:
        r.setdefault("exclusive", True)
    roles.sort(key=lambda r: r["xp"])
    return roles

async def apply_xp_roles(
    guild: discord.Guild, member: discord.Member, guild_data: dict, current_xp: int, announce: bool = True
) -> dict:
    """Applies configured XP-milestone roles, split into two independent
    groups:

    - EXCLUSIVE roles form a tier ladder — only the single highest
      exclusive tier the member currently qualifies for is kept; reaching
      a new exclusive tier removes any other exclusive tier they hold.
      This re-evaluates fresh every time, so a correction that lowers
      someone's XP can demote them the same way it would promote them.
    - STICKY (non-exclusive) roles are granted once their threshold is
      reached and never removed by this function, regardless of what
      other roles someone earns later — e.g. a permanent "crew" role that
      shouldn't disappear just because someone leveled up.

    If announce=True (default) and a milestone-announcement channel is
    configured, posts a public announcement for any newly ADDED role
    (never for removals). Set announce=False for bulk operations
    (imports, retroactive grants, full re-syncs) to avoid flooding the
    channel with a burst of messages — those are better summarized by
    the caller instead.

    Returns {"added": [...], "removed": [...]}, both empty if nothing
    changed or the bot lacks permission/hierarchy to act."""
    xp_roles = get_xp_roles(guild_data)
    if not xp_roles:
        return {"added": [], "removed": []}

    exclusive_roles = [m for m in xp_roles if m["exclusive"]]
    sticky_roles = [m for m in xp_roles if not m["exclusive"]]

    to_add, to_remove = [], []

    # Exclusive ladder: keep only the highest qualifying tier.
    exclusive_qualifying = [m for m in exclusive_roles if current_xp >= m["xp"]]
    if exclusive_qualifying:
        target_mapping = exclusive_qualifying[-1]  # ascending sort, so last = highest qualifying tier
        target_role = guild.get_role(target_mapping["role_id"])
        if target_role and target_role not in member.roles:
            to_add.append(target_role)
        for mapping in exclusive_roles:
            if mapping["role_id"] == target_mapping["role_id"]:
                continue
            role = guild.get_role(mapping["role_id"])
            if role and role in member.roles:
                to_remove.append(role)

    # Sticky roles: add if earned, never remove.
    for mapping in sticky_roles:
        if current_xp >= mapping["xp"]:
            role = guild.get_role(mapping["role_id"])
            if role and role not in member.roles:
                to_add.append(role)

    if not to_add and not to_remove:
        return {"added": [], "removed": []}

    try:
        if to_add:
            await member.add_roles(*to_add, reason="XP milestone reached")
        if to_remove:
            await member.remove_roles(*to_remove, reason="Superseded by a higher XP milestone role")
        if to_add and announce:
            await announce_role_milestone(guild, guild_data, member, to_add)
        return {"added": to_add, "removed": to_remove}
    except (discord.Forbidden, discord.HTTPException):
        # Bot lacks Manage Roles permission, or its role sits below the
        # target role in the server's role hierarchy — fail silently here;
        # the caller can't do much about it mid-flow either way.
        return {"added": [], "removed": []}

async def announce_role_milestone(guild: discord.Guild, guild_data: dict, member: discord.Member, roles: list):
    """Posts a public congratulations message when someone earns a new
    milestone role, if a channel has been configured via /setannouncechannel."""
    channel_id = get_announce_channel_id(guild_data)
    if not channel_id:
        return
    channel = guild.get_channel(channel_id)
    if not channel:
        return
    role_names = ", ".join(r.mention for r in roles)
    embed = discord.Embed(
        title="🎉 Milestone Reached!",
        description=f"{member.mention} just earned {role_names}!",
        color=discord.Color.gold(),
    )
    try:
        await channel.send(embed=embed)
    except discord.HTTPException:
        pass


async def log_admin_action(
    guild: discord.Guild, guild_data: dict, admin: discord.abc.User, action: str,
    target: Optional[str] = None, details: Optional[str] = None,
):
    """Posts a record of a data-affecting admin action to the configured
    log channel, if one is set via /setadminlogchannel. Covers destructive
    or data-altering commands — not routine display toggles, to keep the
    log meaningful rather than noisy. Never raises — a logging failure
    should never block the actual action from completing."""
    channel_id = get_admin_log_channel_id(guild_data)
    if not channel_id:
        return
    channel = guild.get_channel(channel_id)
    if not channel:
        return
    embed = discord.Embed(title="🛡️ Admin Action", color=discord.Color.dark_grey(), timestamp=utcnow())
    embed.add_field(name="Action", value=action, inline=True)
    embed.add_field(name="By", value=admin.mention, inline=True)
    if target:
        embed.add_field(name="Target", value=target, inline=True)
    if details:
        embed.add_field(name="Details", value=details, inline=False)
    try:
        await channel.send(embed=embed)
    except discord.HTTPException:
        pass


def join_leaderboard_lines(lines: list, limit: int = 3900) -> str:
    """Joins leaderboard lines, staying under Discord's 4096-char embed
    description limit. If it would overflow, truncates and notes how many
    entries were cut — better than the whole command failing outright."""
    out, used = [], 0
    for i, line in enumerate(lines):
        needed = len(line) + 1
        if used + needed > limit:
            out.append(f"…and {len(lines) - i} more (enable compact mode in `/settings` to fit more)")
            break
        out.append(line)
        used += needed
    return "\n".join(out)

LEADERBOARD_PAGE_SIZE = 10

class LeaderboardPaginatorView(discord.ui.View):
    """Previous/Next paging through a long leaderboard instead of
    truncating it into a single embed. Used when leaderboard_pagination
    is enabled and the list is long enough to need more than one page."""

    def __init__(self, title: str, lines: list, color: discord.Color, footer: str,
                 extra_field: Optional[tuple] = None, page_size: int = LEADERBOARD_PAGE_SIZE):
        super().__init__(timeout=180)
        self.title = title
        self.lines = lines
        self.color = color
        self.footer = footer
        self.extra_field = extra_field
        self.page_size = page_size
        self.page = 0
        self.max_page = max((len(lines) - 1) // page_size, 0)
        self._update_buttons()

    def _update_buttons(self):
        self.prev_page.disabled = self.page <= 0
        self.next_page.disabled = self.page >= self.max_page

    def build_embed(self) -> discord.Embed:
        start = self.page * self.page_size
        chunk = self.lines[start:start + self.page_size]
        embed = discord.Embed(title=self.title, description="\n".join(chunk) or "No entries.", color=self.color)
        if self.extra_field:
            embed.insert_field_at(0, name=self.extra_field[0], value=self.extra_field[1], inline=False)
        footer = self.footer
        if self.max_page > 0:
            footer += f"  •  Page {self.page + 1}/{self.max_page + 1}"
        embed.set_footer(text=footer)
        return embed

    @discord.ui.button(label="◀ Previous", style=discord.ButtonStyle.secondary)
    async def prev_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.page = max(self.page - 1, 0)
        self._update_buttons()
        await interaction.response.edit_message(embed=self.build_embed(), view=self)

    @discord.ui.button(label="Next ▶", style=discord.ButtonStyle.secondary)
    async def next_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.page = min(self.page + 1, self.max_page)
        self._update_buttons()
        await interaction.response.edit_message(embed=self.build_embed(), view=self)


def build_leaderboard_payload(guild_data: dict, title: str, lines: list, color: discord.Color,
                               footer: str, extra_field: Optional[tuple] = None):
    """Builds the (embed, view) pair for a leaderboard, choosing between
    the paginated and single-truncated-embed presentation based on the
    guild's setting. view is None when pagination isn't used or isn't
    needed (short enough to fit one page). Returns (None, None) if lines
    is empty — the caller should show a "no one tracked yet" message."""
    if not lines:
        return None, None
    if leaderboard_pagination_enabled(guild_data) and len(lines) > LEADERBOARD_PAGE_SIZE:
        view = LeaderboardPaginatorView(title, lines, color, footer, extra_field)
        return view.build_embed(), view
    embed = discord.Embed(title=title, description=join_leaderboard_lines(lines), color=color)
    if extra_field:
        embed.insert_field_at(0, name=extra_field[0], value=extra_field[1], inline=False)
    embed.set_footer(text=footer)
    return embed, None

def create_user(guild_data: dict, user_id: int, xp: int) -> dict:
    """Starts tracking a new user. If the guild has a shared week_anchor set
    (via /setweekprogressall), the user's week is aligned to it instead of
    starting a fresh 7-day window from right now — this keeps late joiners
    on the same weekly clock as everyone else. Their week_start_xp is still
    set to their current XP, since we have no way of knowing what they had
    at the true start of the week.

    baseline_xp is separate from week_start_xp: it's set once, here, and
    never touched by weekly rollovers — it's the reference point for
    all-time total XP gained. It can be corrected later with /setbaseline
    if the very first checkin wasn't actually the true starting point."""
    now = utcnow()
    anchor = guild_data.get("week_anchor")
    week_start = parse_iso(anchor) if anchor else now
    entry = {
        "current_xp": xp,
        "baseline_xp": xp,
        "baseline_time": iso(now),
        "week_start_time": iso(week_start),
        "week_start_xp": xp,
        "checkins": [{"time": iso(now), "xp": xp}],
    }
    guild_data["users"][str(user_id)] = entry
    return entry

def get_baseline(entry: dict) -> tuple:
    """Returns (baseline_xp, baseline_time), auto-migrating older entries
    created before baseline tracking existed. For those, the earliest
    stored checkin is used as a best-effort guess — note this may
    understate total gain if older checkins have already been trimmed
    from history (see MAX_CHECKINS_STORED)."""
    if "baseline_xp" in entry:
        return entry["baseline_xp"], entry.get("baseline_time", entry["checkins"][0]["time"])
    first_checkin = entry["checkins"][0] if entry["checkins"] else {"xp": entry["current_xp"], "time": iso(utcnow())}
    entry["baseline_xp"] = first_checkin["xp"]
    entry["baseline_time"] = first_checkin["time"]
    return entry["baseline_xp"], entry["baseline_time"]

# ---------------------------------------------------------------------------
# Spreadsheet import
# ---------------------------------------------------------------------------

def cell_to_str(v) -> str:
    """Converts an xlsx cell value to a clean string. Whole-number floats
    (e.g. 166615.0, which spreadsheet exports commonly produce even for
    integer-looking numbers) become '166615' instead of '166615.0', which
    would otherwise fail int() parsing downstream."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()

def parse_spreadsheet_rows(filename: str, data: bytes) -> list:
    """Parses a .csv or .xlsx file into a list of dicts, with lowercase,
    stripped column names. Raises ValueError on unsupported/unreadable files."""
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if ext == "csv":
        text = data.decode("utf-8-sig", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        rows = []
        for row in reader:
            clean = {}
            for k, v in row.items():
                if k is None:
                    continue  # extra unnamed columns from a ragged row
                v_str = v if isinstance(v, str) else ""  # missing cells come through as None
                clean[k.strip().lower()] = v_str.strip()
            rows.append(clean)
        return rows

    if ext in ("xlsx", "xlsm"):
        if openpyxl is None:
            raise ValueError(
                "Reading .xlsx files requires the `openpyxl` package, which isn't installed. "
                "Run `pip install openpyxl`, or export your spreadsheet as .csv instead."
            )
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        rows = []
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = [str(c).strip().lower() if c is not None else "" for c in row]
                continue
            row_dict = {header[j]: cell_to_str(v) for j, v in enumerate(row) if j < len(header)}
            if any(v for v in row_dict.values()):  # skip fully blank rows
                rows.append(row_dict)
        return rows

    raise ValueError("Unsupported file type — please upload a .csv or .xlsx file.")

def find_member_for_row(guild: discord.Guild, row: dict) -> tuple:
    """Tries discord_id first (most reliable, always exact). Falls back to
    matching username/display name/global name exactly, then — since many
    servers use a nickname bot that formats nicknames like
    'DiscordName | RobloxUsername' — falls back further to a substring
    match against display names. Returns (member_or_None, match_type),
    where match_type is one of: 'id', 'exact', 'substring',
    'ambiguous_exact', 'ambiguous_substring', 'none'."""
    for id_key in ("discord_id", "id", "user_id"):
        raw_id = row.get(id_key)
        if raw_id:
            cleaned_id = str(raw_id).strip()
            if cleaned_id.endswith(".0"):
                cleaned_id = cleaned_id[:-2]  # tolerates a '...0.0'-style float-ified ID
            try:
                member = guild.get_member(int(cleaned_id))
                if member:
                    return member, "id"
            except ValueError:
                pass

    target = None
    for name_key in ("username", "discord_username", "name", "user", "roblox_username"):
        raw_name = row.get(name_key)
        if raw_name:
            target = raw_name.strip().lstrip("@").lower()
            break

    if not target:
        return None, "none"

    exact = [
        m for m in guild.members
        if m.name.lower() == target
        or m.display_name.lower() == target
        or (getattr(m, "global_name", None) or "").lower() == target
    ]
    if len(exact) == 1:
        return exact[0], "exact"
    if len(exact) > 1:
        return None, "ambiguous_exact"

    # Fallback: the sheet's username appears somewhere inside their nickname —
    # common with bots that format nicknames like "DiscordName | RobloxName".
    substring = [
        m for m in guild.members
        if target in m.display_name.lower() or target in m.name.lower()
    ]
    if len(substring) == 1:
        return substring[0], "substring"
    if len(substring) > 1:
        return None, "ambiguous_substring"

    return None, "none"

def find_xp_for_row(row: dict) -> Optional[int]:
    for xp_key in ("starting_xp", "start_xp", "startxp", "xp"):
        raw_xp = row.get(xp_key)
        if raw_xp not in (None, ""):
            cleaned = str(raw_xp).replace(",", "").strip()
            try:
                return int(cleaned)
            except ValueError:
                try:
                    return int(float(cleaned))  # tolerates '166615.0'-style strings
                except ValueError:
                    return None
    return None

def ensure_current_week(entry: dict) -> bool:
    """Roll the tracking window forward if 7+ days have elapsed since
    week_start_time. Uses the most recent known XP as the new week's
    starting point (an approximation, since we only know XP at checkin
    times, not at the exact week boundary)."""
    now = utcnow()
    week_start = parse_iso(entry["week_start_time"])
    rolled = False
    while now - week_start >= timedelta(days=7):
        week_start = week_start + timedelta(days=7)
        entry["week_start_time"] = iso(week_start)
        entry["week_start_xp"] = entry["current_xp"]
        rolled = True
    return rolled

def record_checkin(entry: dict, xp: int) -> Optional[dict]:
    """Records a new checkin, returns the previous checkin (or None)."""
    ensure_current_week(entry)
    now = utcnow()
    last = entry["checkins"][-1] if entry["checkins"] else None
    entry["current_xp"] = xp
    entry["checkins"].append({"time": iso(now), "xp": xp})
    if len(entry["checkins"]) > MAX_CHECKINS_STORED:
        entry["checkins"] = entry["checkins"][-MAX_CHECKINS_STORED:]
    return last

def compute_stats(entry: dict, requirement: int, prorate_threshold_hours: float = DEFAULT_PRORATE_THRESHOLD_HOURS) -> dict:
    ensure_current_week(entry)
    now = utcnow()
    week_start = parse_iso(entry["week_start_time"])
    elapsed_days = max((now - week_start).total_seconds() / 86400, 1 / 24)  # floor at 1 hour
    gained_this_week = entry["current_xp"] - entry["week_start_xp"]
    rate_per_day = gained_this_week / elapsed_days
    days_left = max(7 - elapsed_days, 0)
    projected_total = gained_this_week + rate_per_day * days_left

    # If this user's tracking started partway through the CURRENT week (e.g. they
    # joined a server that already had a shared week in progress), it's unfair to
    # hold them to the full flat requirement — they never had the full 7 days.
    # Prorate the requirement to the fraction of the week they actually had,
    # based on time from their first-ever checkin to this week's end. Only
    # applies to that first partial week; once they've rolled into a full week,
    # baseline_time falls before week_start and this has no effect.
    #
    # prorate_threshold_hours controls how late is "late enough" to matter:
    # proration only kicks in if the hours available to them at their first
    # checkin were at or below this threshold. Default is the full week (168h),
    # meaning any lateness at all triggers it; a server can raise the bar so
    # only genuinely late joiners (e.g. <24h left) get prorated.
    baseline_xp, baseline_time = get_baseline(entry)
    baseline_dt = parse_iso(baseline_time)
    week_end = week_start + timedelta(days=7)
    available_hours_at_join = (week_end - baseline_dt).total_seconds() / 3600
    is_prorated = baseline_dt > week_start and available_hours_at_join <= prorate_threshold_hours
    if is_prorated:
        fraction_available = max(min(available_hours_at_join / (7 * 24), 1.0), 0.0)
        effective_requirement = requirement * fraction_available
    else:
        effective_requirement = requirement

    remaining_needed = max(effective_requirement - gained_this_week, 0)
    if days_left > 0:
        required_rate_per_day = remaining_needed / days_left
    else:
        required_rate_per_day = float("inf") if remaining_needed > 0 else 0
    on_track = projected_total >= effective_requirement

    MIN_INTERVAL_HOURS = 1 / 6  # 10 minutes — below this, extrapolating to an hourly rate is too noisy to trust
    recent_rate_per_hour = None
    recent_interval_hours = None
    if len(entry["checkins"]) >= 2:
        c1, c2 = entry["checkins"][-2], entry["checkins"][-1]
        dt_hours = (parse_iso(c2["time"]) - parse_iso(c1["time"])).total_seconds() / 3600
        if dt_hours >= MIN_INTERVAL_HOURS:
            recent_rate_per_hour = (c2["xp"] - c1["xp"]) / dt_hours
            recent_interval_hours = dt_hours

    checkins_this_week = sum(1 for c in entry["checkins"] if parse_iso(c["time"]) >= week_start)

    return {
        "gained_this_week": gained_this_week,
        "elapsed_days": elapsed_days,
        "days_left": days_left,
        "rate_per_day": rate_per_day,
        "projected_total": projected_total,
        "remaining_needed": remaining_needed,
        "required_rate_per_day": required_rate_per_day,
        "on_track": on_track,
        "recent_rate_per_hour": recent_rate_per_hour,
        "recent_interval_hours": recent_interval_hours,
        "checkins_this_week": checkins_this_week,
        "is_prorated": is_prorated,
        "effective_requirement": effective_requirement,
    }

# ---------------------------------------------------------------------------
# Embeds
# ---------------------------------------------------------------------------

def build_status_embed(member: discord.abc.User, entry: dict, stats: dict, requirement: int,
                        extra_field: Optional[tuple] = None, show_recent_rate: bool = True) -> discord.Embed:
    color = discord.Color.green() if stats["on_track"] else discord.Color.orange()
    embed = discord.Embed(title=f"\U0001F4C8 XP Status \u2014 {member.display_name}", color=color)
    embed.add_field(name="📅 Week Runs", value=week_range_str(parse_iso(entry["week_start_time"])), inline=False)

    progress_target = stats["effective_requirement"] if stats["is_prorated"] else requirement
    bar = build_progress_bar(stats["gained_this_week"], progress_target)
    embed.add_field(name="Progress Toward Requirement", value=bar, inline=False)

    embed.add_field(name="Current XP", value=f"{entry['current_xp']:,}", inline=True)
    embed.add_field(name="Gained This Week", value=f"{stats['gained_this_week']:,}", inline=True)
    if stats["is_prorated"]:
        embed.add_field(
            name="Weekly Requirement",
            value=f"{stats['effective_requirement']:,.0f} (prorated)\nFull: {requirement:,}",
            inline=True,
        )
    else:
        embed.add_field(name="Weekly Requirement", value=f"{requirement:,}", inline=True)

    embed.add_field(name="Avg Rate", value=f"{stats['rate_per_day']:,.0f} XP/day", inline=True)
    if show_recent_rate and stats["recent_rate_per_hour"] is not None:
        interval_str = format_duration(stats["recent_interval_hours"])
        embed.add_field(
            name="Recent Rate",
            value=f"{stats['recent_rate_per_hour']:,.1f} XP/hr\n(based on last {interval_str})",
            inline=True,
        )
    embed.add_field(name="Days Left in Week", value=format_duration(stats["days_left"] * 24), inline=True)

    embed.add_field(name="Projected Weekly Total", value=f"{stats['projected_total']:,.0f}", inline=True)
    if stats["remaining_needed"] > 0:
        req_rate = stats["required_rate_per_day"]
        req_str = "∞" if req_rate == float("inf") else f"{req_rate:,.0f}/day"
        embed.add_field(name="Still Needed", value=f"{stats['remaining_needed']:,.0f} XP ({req_str})", inline=True)
    embed.add_field(name="On Track?", value="✅ Yes" if stats["on_track"] else "⚠️ Behind pace", inline=True)

    if stats["is_prorated"]:
        embed.add_field(
            name="ℹ️ Prorated First Week",
            value=(
                f"This user's tracking started partway through the current week, so their "
                f"requirement was scaled down to **{stats['effective_requirement']:,.0f}** — "
                f"the fair share of {requirement:,} for the time they actually had. "
                f"Next week uses the full requirement."
            ),
            inline=False,
        )

    if stats["checkins_this_week"] <= 1:
        embed.add_field(
            name="⚠️ Low Data Confidence",
            value=(
                "Only 1 checkin recorded this week, so this number is entirely "
                "self-reported with nothing to cross-check it against. Encourage "
                "more frequent checkins (e.g. every day or two) for a trustworthy pace."
            ),
            inline=False,
        )

    if extra_field:
        embed.add_field(name=extra_field[0], value=extra_field[1], inline=False)

    return embed

def build_settings_embed(guild_data: dict) -> discord.Embed:
    embed = discord.Embed(title="⚙️ XP Tracker Settings", color=discord.Color.blurple())
    embed.add_field(name="Weekly Requirement", value=f"{guild_data['requirement']:,} XP", inline=True)

    anchor = guild_data.get("week_anchor")
    if anchor:
        anchor_ts = int(parse_iso(anchor).timestamp())
        anchor_str = f"<t:{anchor_ts}:R>"
    else:
        anchor_str = "Not set — new users start their own week"
    embed.add_field(name="Shared Week Anchor", value=anchor_str, inline=True)
    embed.add_field(name="Tracked Users", value=str(len(guild_data["users"])), inline=True)
    embed.add_field(
        name="Recent Rate Field",
        value="Shown" if show_recent_rate_enabled(guild_data) else "Hidden",
        inline=True,
    )
    embed.add_field(
        name="Leaderboard Style",
        value="Compact" if compact_leaderboard_enabled(guild_data) else "Full detail",
        inline=True,
    )
    threshold = get_prorate_threshold_hours(guild_data)
    if threshold >= 168:
        threshold_str = "Any lateness (default)"
    elif threshold <= 0:
        threshold_str = "Off"
    else:
        threshold_str = f"≤ {format_duration(threshold)} left"
    embed.add_field(name="Prorate Threshold", value=threshold_str, inline=True)
    embed.add_field(name="XP Milestone Roles", value=str(len(get_xp_roles(guild_data))), inline=True)

    restricted = get_restricted_channel_id(guild_data)
    embed.add_field(
        name="Tracking Channel",
        value=f"<#{restricted}>" if restricted else "Any channel (unrestricted)",
        inline=True,
    )
    announce = get_announce_channel_id(guild_data)
    embed.add_field(
        name="Milestone Announcements",
        value=f"<#{announce}>" if announce else "Off",
        inline=True,
    )
    weekly_post = get_weekly_post_channel_id(guild_data)
    embed.add_field(
        name="Weekly Auto-Post",
        value=f"<#{weekly_post}>" if weekly_post else "Off",
        inline=True,
    )
    inactivity_ch = get_inactivity_channel_id(guild_data)
    if inactivity_ch:
        scope = "+ behind pace" if inactivity_includes_behind_pace(guild_data) else "zero checkins only"
        inactivity_val = f"<#{inactivity_ch}> ({format_duration(get_inactivity_threshold_hours(guild_data))} before end, {scope})"
    else:
        inactivity_val = "Off"
    embed.add_field(name="Inactivity Pings", value=inactivity_val, inline=True)
    embed.add_field(
        name="Leaderboard Pagination",
        value="On" if leaderboard_pagination_enabled(guild_data) else "Off",
        inline=True,
    )
    admin_log_ch = get_admin_log_channel_id(guild_data)
    embed.add_field(name="Admin Action Log", value=f"<#{admin_log_ch}>" if admin_log_ch else "Off", inline=True)

    embed.set_footer(text="Use the buttons below to change settings. Panel expires after 5 minutes of inactivity.")
    return embed

def build_progress_chart(entry: dict, display_name: str) -> Optional[io.BytesIO]:
    """Renders a simple XP-over-time line chart from a user's checkin
    history as a PNG. Returns None if matplotlib isn't installed or
    there's fewer than 2 checkins to plot."""
    if plt is None or len(entry["checkins"]) < 2:
        return None

    times = [parse_iso(c["time"]) for c in entry["checkins"]]
    xps = [c["xp"] for c in entry["checkins"]]

    fig, ax = plt.subplots(figsize=(8, 4.5), dpi=120)
    ax.plot(times, xps, marker="o", markersize=4, linewidth=2, color="#5865F2")
    ax.fill_between(times, xps, min(xps), alpha=0.08, color="#5865F2")

    ax.set_title(f"XP Progress — {display_name}", fontsize=13, fontweight="bold", color="#2C2F33")
    ax.set_ylabel("Total XP", fontsize=10)
    ax.yaxis.set_major_formatter(lambda x, pos: f"{x:,.0f}")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %d"))
    fig.autofmt_xdate(rotation=30)
    ax.grid(True, alpha=0.25, linestyle="--")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.tight_layout()

    buf = io.BytesIO()
    fig.savefig(buf, format="png", transparent=False, facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf

# ---------------------------------------------------------------------------
# Interactive settings panel (buttons, modals, select menus)
# ---------------------------------------------------------------------------

class RequirementModal(discord.ui.Modal, title="Set Weekly XP Requirement"):
    requirement = discord.ui.TextInput(
        label="New weekly XP requirement", placeholder="e.g. 20000", max_length=12
    )

    def __init__(self, guild_id: int):
        super().__init__()
        self.guild_id = guild_id

    async def on_submit(self, interaction: discord.Interaction):
        try:
            xp = int(self.requirement.value.replace(",", "").strip())
        except ValueError:
            await interaction.response.send_message("Please enter a whole number.", ephemeral=True)
            return
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        guild_data["requirement"] = xp
        save_data(data)
        await interaction.response.send_message(f"Weekly requirement set to **{xp:,} XP**.", ephemeral=True)


def parse_dhm_fields(days_str: str, hours_str: str, minutes_str: str) -> timedelta:
    """Parses three optional text fields (blank = 0) into a timedelta,
    raising ValueError with a user-facing message if anything's invalid."""
    def to_int(s: str, label: str) -> int:
        s = (s or "").strip()
        if not s:
            return 0
        try:
            return int(s)
        except ValueError:
            raise ValueError(f"{label} must be a whole number.")

    days = to_int(days_str, "Days")
    hours = to_int(hours_str, "Hours")
    minutes = to_int(minutes_str, "Minutes")
    delta = timedelta(days=days, hours=hours, minutes=minutes)
    if delta < timedelta(0) or delta > timedelta(days=7):
        raise ValueError("Total time must be between 0 and 7 days.")
    return delta


class ProrateThresholdModal(discord.ui.Modal, title="Set Prorate Threshold"):
    days = discord.ui.TextInput(label="Days remaining (0-7)", placeholder="e.g. 1", required=False, max_length=3)
    hours = discord.ui.TextInput(label="Plus hours (0-23)", placeholder="e.g. 0", required=False, max_length=3)
    minutes = discord.ui.TextInput(label="Plus minutes (0-59)", placeholder="e.g. 0", required=False, max_length=3)

    def __init__(self, guild_id: int):
        super().__init__()
        self.guild_id = guild_id

    async def on_submit(self, interaction: discord.Interaction):
        try:
            delta = parse_dhm_fields(self.days.value, self.hours.value, self.minutes.value)
        except ValueError as e:
            await interaction.response.send_message(str(e), ephemeral=True)
            return

        threshold_hours = delta.total_seconds() / 3600
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        guild_data["prorate_threshold_hours"] = threshold_hours
        save_data(data)

        if threshold_hours >= 168:
            note = "Any late join at all now gets a prorated requirement."
        elif threshold_hours <= 0:
            note = "Proration is now effectively off."
        else:
            note = f"Only joins with **{format_duration(threshold_hours)} or less** remaining will be prorated."
        await interaction.response.send_message(
            f"Prorate threshold set to {format_duration(threshold_hours)}. {note}", ephemeral=True
        )


class XPThresholdModal(discord.ui.Modal, title="Set XP Threshold"):
    xp_threshold = discord.ui.TextInput(label="XP required for this role", placeholder="e.g. 75000", max_length=12)
    sticky = discord.ui.TextInput(
        label="Permanent role? (yes/no)",
        placeholder="no = tier ladder (removed by higher tiers), yes = kept forever",
        required=False,
        max_length=3,
    )

    def __init__(self, guild_id: int, role: discord.Role):
        super().__init__()
        self.guild_id = guild_id
        self.role = role

    async def on_submit(self, interaction: discord.Interaction):
        try:
            xp = int(self.xp_threshold.value.replace(",", "").strip())
        except ValueError:
            await interaction.response.send_message("Please enter a whole number.", ephemeral=True)
            return
        if xp < 0:
            await interaction.response.send_message("XP must be a positive number.", ephemeral=True)
            return

        is_sticky = self.sticky.value.strip().lower() in ("yes", "y", "true")
        exclusive = not is_sticky

        bot_member = interaction.guild.me
        if self.role >= bot_member.top_role:
            await interaction.response.send_message(
                f"⚠️ I can't assign {self.role.mention} — it's at or above my own highest role. "
                f"Move my role above it in Server Settings → Roles, then try again.",
                ephemeral=True,
            )
            return

        await interaction.response.defer(thinking=True, ephemeral=True)

        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        xp_roles = guild_data.setdefault("xp_roles", [])
        xp_roles[:] = [r for r in xp_roles if r["role_id"] != self.role.id]
        xp_roles.append({"xp": xp, "role_id": self.role.id, "exclusive": exclusive})
        xp_roles.sort(key=lambda r: r["xp"])

        granted = 0
        for uid, entry in guild_data["users"].items():
            if entry["current_xp"] >= xp:
                member = interaction.guild.get_member(int(uid))
                if member and self.role not in member.roles:
                    result = await apply_xp_roles(interaction.guild, member, guild_data, entry["current_xp"], announce=False)
                    if result["added"]:
                        granted += 1
        save_data(data)

        kind = "permanent (kept even after higher tiers)" if is_sticky else "part of the tier ladder"
        note = f" Retroactively granted to **{granted}** member(s) who already qualify." if granted else ""
        await interaction.followup.send(
            f"Members reaching **{xp:,} XP** will now automatically receive {self.role.mention} — {kind}.{note}",
            ephemeral=True,
        )


class AddXPRoleSelectView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=120)
        self.guild_id = guild_id

    @discord.ui.select(cls=discord.ui.RoleSelect, placeholder="Choose a role to assign at an XP threshold")
    async def select_role(self, interaction: discord.Interaction, select: discord.ui.RoleSelect):
        role = select.values[0]
        bot_member = interaction.guild.me
        if not bot_member.guild_permissions.manage_roles:
            await interaction.response.send_message(
                "⚠️ I don't have the **Manage Roles** permission in this server yet.", ephemeral=True
            )
            return
        await interaction.response.send_modal(XPThresholdModal(self.guild_id, role))


class RemoveXPRoleSelectView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=120)
        self.guild_id = guild_id

    @discord.ui.select(cls=discord.ui.RoleSelect, placeholder="Choose a role to stop auto-assigning")
    async def select_role(self, interaction: discord.Interaction, select: discord.ui.RoleSelect):
        role = select.values[0]
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        xp_roles = guild_data.setdefault("xp_roles", [])
        before = len(xp_roles)
        xp_roles[:] = [r for r in xp_roles if r["role_id"] != role.id]
        save_data(data)
        if len(xp_roles) < before:
            await interaction.response.send_message(
                f"{role.mention} will no longer be auto-assigned. Members who already have it keep it.",
                ephemeral=True,
            )
        else:
            await interaction.response.send_message("That role wasn't configured as an XP milestone.", ephemeral=True)


class SetTrackingChannelView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=120)
        self.guild_id = guild_id

    @discord.ui.select(cls=discord.ui.ChannelSelect, channel_types=[discord.ChannelType.text],
                        placeholder="Choose the channel for /checkin, /status, and leaderboards")
    async def select_channel(self, interaction: discord.Interaction, select: discord.ui.ChannelSelect):
        channel = select.values[0]
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        guild_data["restricted_channel_id"] = channel.id
        save_data(data)
        await interaction.response.send_message(
            f"Tracking commands are now restricted to {channel.mention}. Use `/clearchannel` to remove this restriction.",
            ephemeral=True,
        )


class SetAnnounceChannelView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=120)
        self.guild_id = guild_id

    @discord.ui.select(cls=discord.ui.ChannelSelect, channel_types=[discord.ChannelType.text],
                        placeholder="Choose the channel for milestone announcements")
    async def select_channel(self, interaction: discord.Interaction, select: discord.ui.ChannelSelect):
        channel = select.values[0]
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        guild_data["announce_channel_id"] = channel.id
        save_data(data)
        await interaction.response.send_message(
            f"Milestone announcements will now post in {channel.mention}. Use `/clearannouncechannel` to turn this off.",
            ephemeral=True,
        )


class SetWeeklyPostChannelView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=120)
        self.guild_id = guild_id

    @discord.ui.select(cls=discord.ui.ChannelSelect, channel_types=[discord.ChannelType.text],
                        placeholder="Choose the channel for the auto-posted weekly leaderboard")
    async def select_channel(self, interaction: discord.Interaction, select: discord.ui.ChannelSelect):
        channel = select.values[0]
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        if not guild_data.get("week_anchor"):
            await interaction.response.send_message(
                "This requires a shared server week first — run `/setweekprogressall` at least once.",
                ephemeral=True,
            )
            return
        guild_data["weekly_post_channel_id"] = channel.id
        save_data(data)
        await interaction.response.send_message(
            f"The weekly leaderboard will now auto-post in {channel.mention} shortly before each week ends. "
            f"Use `/clearweeklypost` to turn this off.",
            ephemeral=True,
        )


class SetInactivityChannelView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=120)
        self.guild_id = guild_id

    @discord.ui.select(cls=discord.ui.ChannelSelect, channel_types=[discord.ChannelType.text],
                        placeholder="Choose the channel for inactivity reminder pings")
    async def select_channel(self, interaction: discord.Interaction, select: discord.ui.ChannelSelect):
        channel = select.values[0]
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        if not guild_data.get("week_anchor"):
            await interaction.response.send_message(
                "This requires a shared server week first — run `/setweekprogressall` at least once.",
                ephemeral=True,
            )
            return
        guild_data["inactivity_channel_id"] = channel.id
        save_data(data)
        threshold = get_inactivity_threshold_hours(guild_data)
        await interaction.response.send_message(
            f"Inactivity reminders will now post in {channel.mention} once **{format_duration(threshold)}** remain "
            f"in the week. Adjust timing with `/setinactivitythreshold`, or turn off with `/clearinactivitychannel`.",
            ephemeral=True,
        )


class SetAdminLogChannelView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=120)
        self.guild_id = guild_id

    @discord.ui.select(cls=discord.ui.ChannelSelect, channel_types=[discord.ChannelType.text],
                        placeholder="Choose the channel for admin action logs")
    async def select_channel(self, interaction: discord.Interaction, select: discord.ui.ChannelSelect):
        channel = select.values[0]
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        guild_data["admin_log_channel_id"] = channel.id
        save_data(data)
        await interaction.response.send_message(
            f"Admin actions (resets, removals, imports, requirement changes, etc.) will now be logged "
            f"in {channel.mention}. Turn off with `/clearadminlogchannel`.",
            ephemeral=True,
        )


class SyncWeekModal(discord.ui.Modal, title="Sync Week Progress (Everyone)"):
    days = discord.ui.TextInput(label="Days (0-7)", placeholder="e.g. 3", required=False, max_length=3)
    hours = discord.ui.TextInput(label="Hours (0-23)", placeholder="e.g. 12", required=False, max_length=3)
    minutes = discord.ui.TextInput(label="Minutes (0-59)", placeholder="e.g. 30", required=False, max_length=3)
    reset_gains = discord.ui.TextInput(
        label="Reset everyone's gained-so-far? (yes/no)", placeholder="no", required=False, max_length=3
    )

    def __init__(self, guild_id: int):
        super().__init__()
        self.guild_id = guild_id

    async def on_submit(self, interaction: discord.Interaction):
        try:
            delta = parse_dhm_fields(self.days.value, self.hours.value, self.minutes.value)
        except ValueError as e:
            await interaction.response.send_message(str(e), ephemeral=True)
            return

        reset = self.reset_gains.value.strip().lower() in ("yes", "y", "true")

        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        new_week_start = utcnow() - delta
        guild_data["week_anchor"] = iso(new_week_start)

        count = 0
        for entry in guild_data["users"].values():
            entry["week_start_time"] = iso(new_week_start)
            if reset:
                entry["week_start_xp"] = entry["current_xp"]
            count += 1
        save_data(data)

        gains_note = " Gained-so-far was reset for everyone." if reset else " Existing gained-so-far totals were kept."
        await interaction.response.send_message(
            f"Set week progress to **{format_timedelta(delta)} elapsed** for **{count}** user(s)."
            f"{gains_note}",
            ephemeral=True,
        )


class UserWeekProgressModal(discord.ui.Modal, title="Set Week Progress for User"):
    days = discord.ui.TextInput(label="Days (0-7)", placeholder="e.g. 3", required=False, max_length=3)
    hours = discord.ui.TextInput(label="Hours (0-23)", placeholder="e.g. 12", required=False, max_length=3)
    minutes = discord.ui.TextInput(label="Minutes (0-59)", placeholder="e.g. 30", required=False, max_length=3)
    week_start_xp = discord.ui.TextInput(
        label="Override week-start XP (optional)", required=False, placeholder="leave blank to keep existing"
    )

    def __init__(self, guild_id: int, member: discord.Member):
        super().__init__()
        self.guild_id = guild_id
        self.member = member

    async def on_submit(self, interaction: discord.Interaction):
        try:
            delta = parse_dhm_fields(self.days.value, self.hours.value, self.minutes.value)
        except ValueError as e:
            await interaction.response.send_message(str(e), ephemeral=True)
            return

        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        entry = guild_data["users"].get(str(self.member.id))
        if entry is None:
            await interaction.response.send_message("That user has no tracking data.", ephemeral=True)
            return

        entry["week_start_time"] = iso(utcnow() - delta)
        if self.week_start_xp.value.strip():
            try:
                entry["week_start_xp"] = int(self.week_start_xp.value.replace(",", "").strip())
            except ValueError:
                await interaction.response.send_message("Week-start XP must be a whole number.", ephemeral=True)
                return
        save_data(data)
        await interaction.response.send_message(
            f"Set week progress for **{self.member.display_name}** to "
            f"**{format_timedelta(delta)} elapsed**.",
            ephemeral=True,
        )


class UserBaselineModal(discord.ui.Modal, title="Set All-Time Starting XP"):
    starting_xp = discord.ui.TextInput(label="True starting XP", placeholder="e.g. 50000", max_length=12)

    def __init__(self, guild_id: int, member: discord.Member):
        super().__init__()
        self.guild_id = guild_id
        self.member = member

    async def on_submit(self, interaction: discord.Interaction):
        try:
            xp = int(self.starting_xp.value.replace(",", "").strip())
        except ValueError:
            await interaction.response.send_message("Please enter a whole number.", ephemeral=True)
            return

        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        entry = guild_data["users"].get(str(self.member.id))
        if entry is None:
            await interaction.response.send_message("That user has no tracking data.", ephemeral=True)
            return

        entry["baseline_xp"] = xp
        entry["baseline_time"] = iso(utcnow())
        save_data(data)
        total_gained = entry["current_xp"] - xp
        await interaction.response.send_message(
            f"Set **{self.member.display_name}**'s starting XP to **{xp:,}**. "
            f"Total gained is now **{total_gained:,} XP**.",
            ephemeral=True,
        )


class ConfirmRemoveAllView(discord.ui.View):
    """Two-step confirmation for wiping every tracked user in a server —
    destructive and irreversible, so it never fires on a single click."""

    def __init__(self, guild_id: int, user_count: int):
        super().__init__(timeout=60)
        self.guild_id = guild_id
        self.user_count = user_count
        self.message: Optional[discord.Message] = None

    @discord.ui.button(label="Confirm — Delete Everyone", style=discord.ButtonStyle.danger, emoji="🗑️")
    async def confirm(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        removed = len(guild_data["users"])
        guild_data["users"] = {}
        save_data(data)
        for child in self.children:
            child.disabled = True
        await interaction.response.edit_message(
            content=f"✅ Removed tracking data for **{removed}** user(s). Everyone starts fresh with their next `/checkin`.",
            view=self,
        )
        await log_admin_action(
            interaction.guild, guild_data, interaction.user, "Removed ALL users",
            details=f"{removed} user(s) removed.",
        )

    @discord.ui.button(label="Cancel", style=discord.ButtonStyle.secondary, emoji="✖️")
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        for child in self.children:
            child.disabled = True
        await interaction.response.edit_message(content="Cancelled — no data was removed.", view=self)

    async def on_timeout(self):
        for child in self.children:
            child.disabled = True
        if self.message:
            try:
                await self.message.edit(content="Confirmation timed out — no data was removed.", view=self)
            except discord.HTTPException:
                pass


class ConfirmRemoveStaleView(discord.ui.View):
    """Two-step confirmation for wiping tracking data belonging to users
    who are no longer in the server."""

    def __init__(self, guild_id: int, stale_uids: list):
        super().__init__(timeout=60)
        self.guild_id = guild_id
        self.stale_uids = stale_uids
        self.message: Optional[discord.Message] = None

    @discord.ui.button(label="Confirm — Remove Departed Members", style=discord.ButtonStyle.danger, emoji="🗑️")
    async def confirm(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        removed = 0
        for uid in self.stale_uids:
            if uid in guild_data["users"]:
                del guild_data["users"][uid]
                removed += 1
        save_data(data)
        for child in self.children:
            child.disabled = True
        await interaction.response.edit_message(
            content=f"✅ Removed tracking data for **{removed}** member(s) no longer in the server.",
            view=self,
        )
        await log_admin_action(
            interaction.guild, guild_data, interaction.user, "Removed departed members",
            details=f"{removed} member(s) removed.",
        )

    @discord.ui.button(label="Cancel", style=discord.ButtonStyle.secondary, emoji="✖️")
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        for child in self.children:
            child.disabled = True
        await interaction.response.edit_message(content="Cancelled — no data was removed.", view=self)

    async def on_timeout(self):
        for child in self.children:
            child.disabled = True
        if self.message:
            try:
                await self.message.edit(content="Confirmation timed out — no data was removed.", view=self)
            except discord.HTTPException:
                pass


class ConfirmUndoImportView(discord.ui.View):
    """Two-step confirmation for reverting an entire /importxp batch back
    to how things were right before it ran."""

    def __init__(self, guild_id: int):
        super().__init__(timeout=60)
        self.guild_id = guild_id
        self.message: Optional[discord.Message] = None

    @discord.ui.button(label="Confirm — Undo Import", style=discord.ButtonStyle.danger, emoji="↩️")
    async def confirm(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        snapshot = guild_data.get("last_import_snapshot")
        for child in self.children:
            child.disabled = True

        if not snapshot:
            await interaction.response.edit_message(content="Nothing to undo — no import snapshot found.", view=self)
            return

        guild_data["users"] = snapshot["users"]
        guild_data["last_import_snapshot"] = None  # one-time undo — can't stack undos of undos
        save_data(data)
        await interaction.response.edit_message(
            content="✅ Reverted to how tracking data looked before that import ran.",
            view=self,
        )
        await log_admin_action(interaction.guild, guild_data, interaction.user, "Undid last import")

    @discord.ui.button(label="Cancel", style=discord.ButtonStyle.secondary, emoji="✖️")
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        for child in self.children:
            child.disabled = True
        await interaction.response.edit_message(content="Cancelled — nothing was changed.", view=self)

    async def on_timeout(self):
        for child in self.children:
            child.disabled = True
        if self.message:
            try:
                await self.message.edit(content="Confirmation timed out — nothing was changed.", view=self)
            except discord.HTTPException:
                pass


class ConfirmRestoreView(discord.ui.View):
    """Two-step confirmation for restoring a full /backup — replaces
    EVERY setting and every tracked user's data for this server."""

    def __init__(self, guild_id: int, restored_guild_data: dict):
        super().__init__(timeout=60)
        self.guild_id = guild_id
        self.restored_guild_data = restored_guild_data
        self.message: Optional[discord.Message] = None

    @discord.ui.button(label="Confirm — Restore Backup", style=discord.ButtonStyle.danger, emoji="♻️")
    async def confirm(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        normalized = normalize_guild_data(self.restored_guild_data)
        data["guilds"][str(self.guild_id)] = normalized
        save_data(data)
        for child in self.children:
            child.disabled = True
        await interaction.response.edit_message(
            content=f"✅ Backup restored — {len(normalized['users'])} tracked user(s), all settings replaced.",
            view=self,
        )
        await log_admin_action(
            interaction.guild, normalized, interaction.user, "Restored full backup",
            details=f"{len(normalized['users'])} tracked user(s) after restore.",
        )

    @discord.ui.button(label="Cancel", style=discord.ButtonStyle.secondary, emoji="✖️")
    async def cancel(self, interaction: discord.Interaction, button: discord.ui.Button):
        for child in self.children:
            child.disabled = True
        await interaction.response.edit_message(content="Cancelled — nothing was changed.", view=self)

    async def on_timeout(self):
        for child in self.children:
            child.disabled = True
        if self.message:
            try:
                await self.message.edit(content="Confirmation timed out — nothing was changed.", view=self)
            except discord.HTTPException:
                pass


class UserActionView(discord.ui.View):
    """Shown after picking a user from UserSelectView — one-click actions for that user."""

    def __init__(self, guild_id: int, member: discord.Member):
        super().__init__(timeout=120)
        self.guild_id = guild_id
        self.member = member

    @discord.ui.button(label="Set Week Progress", style=discord.ButtonStyle.primary, emoji="🕒")
    async def set_progress(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(UserWeekProgressModal(self.guild_id, self.member))

    @discord.ui.button(label="Set Starting XP", style=discord.ButtonStyle.primary, emoji="🏁")
    async def set_baseline(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(UserBaselineModal(self.guild_id, self.member))

    @discord.ui.button(label="Reset Week", style=discord.ButtonStyle.secondary, emoji="🔁")
    async def reset_week(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        entry = guild_data["users"].get(str(self.member.id))
        if entry is None:
            await interaction.response.send_message("That user has no tracking data.", ephemeral=True)
            return
        entry["week_start_time"] = iso(utcnow())
        entry["week_start_xp"] = entry["current_xp"]
        save_data(data)
        await interaction.response.send_message(f"Weekly window reset for {self.member.display_name}.", ephemeral=True)

    @discord.ui.button(label="Undo Last Checkin", style=discord.ButtonStyle.secondary, emoji="↩️")
    async def undo_checkin(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        entry = guild_data["users"].get(str(self.member.id))
        if entry is None or not entry["checkins"]:
            await interaction.response.send_message("No checkins to undo.", ephemeral=True)
            return
        if len(entry["checkins"]) == 1:
            del guild_data["users"][str(self.member.id)]
            save_data(data)
            await interaction.response.send_message(
                f"Removed {self.member.display_name}'s only checkin — tracking reset.", ephemeral=True
            )
            return
        removed = entry["checkins"].pop()
        entry["current_xp"] = entry["checkins"][-1]["xp"]
        save_data(data)
        await interaction.response.send_message(
            f"Undid checkin of **{removed['xp']:,} XP** for {self.member.display_name}.", ephemeral=True
        )

    @discord.ui.button(label="Full Reset", style=discord.ButtonStyle.danger, emoji="♻️")
    async def full_reset(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        entry = guild_data["users"].get(str(self.member.id))
        if entry is None:
            await interaction.response.send_message("That user has no tracking data.", ephemeral=True)
            return
        now = utcnow()
        current_xp = entry["current_xp"]
        entry["baseline_xp"] = current_xp
        entry["baseline_time"] = iso(now)
        entry["week_start_time"] = iso(now)
        entry["week_start_xp"] = current_xp
        entry["checkins"] = [{"time": iso(now), "xp": current_xp}]
        save_data(data)
        await interaction.response.send_message(
            f"Fully reset {self.member.display_name} — weekly and all-time totals now start from "
            f"their current XP of **{current_xp:,}**. History was cleared.",
            ephemeral=True,
        )

    @discord.ui.button(label="Remove User", style=discord.ButtonStyle.danger, emoji="🗑️")
    async def remove_user(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        if str(self.member.id) in guild_data["users"]:
            del guild_data["users"][str(self.member.id)]
            save_data(data)
            await interaction.response.send_message(
                f"Removed all tracking data for {self.member.display_name}.", ephemeral=True
            )
        else:
            await interaction.response.send_message("No data found for that user.", ephemeral=True)


class UserSelectView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=120)
        self.guild_id = guild_id

    @discord.ui.select(cls=discord.ui.UserSelect, placeholder="Choose a user to manage")
    async def select_user(self, interaction: discord.Interaction, select: discord.ui.UserSelect):
        member = select.values[0]
        view = UserActionView(self.guild_id, member)
        await interaction.response.edit_message(
            content=f"Managing **{member.display_name}** — choose an action:", view=view
        )


HELP_CATEGORIES = [
    ("tracking", "Tracking", "📝", "Checking in and viewing progress", [
        ("/checkin xp:<number> [user]", "Record current XP — yours, or (admin) someone else's"),
        ("/status [user]", "Weekly progress, rate, and projection for yourself or another user"),
        ("/undo [user]", "Remove the most recent checkin"),
        ("/history [user]", "Last 10 checkins for a user"),
        ("/progresschart [user]", "XP-over-time line chart for a user"),
    ]),
    ("leaderboards", "Leaderboards", "🏆", "Comparing progress across the crew", [
        ("/weeklyleaderboard", "This week's XP ranking"),
        ("/totalleaderboard", "All-time total XP ranking (includes starting point)"),
        ("/crewtotals", "Combined crew-wide totals, including departed members"),
        ("/listxproles", "Show configured XP milestone roles"),
    ]),
    ("admin_core", "Admin: Requirement & Week", "⏱️", "Weekly goal, week sync, proration", [
        ("/settings", "Open the interactive settings panel"),
        ("/setrequirement xp:<number>", "Change the weekly XP goal"),
        ("/setweekprogress user: days: hours: minutes:", "Backdate one user's week progress"),
        ("/setweekprogressall days: hours: minutes:", "Backdate everyone's week progress at once"),
        ("/resetweek user:<@user>", "Restart one user's weekly window right now"),
        ("/setproratethreshold days: hours: minutes:", "How late is \"late enough\" to prorate a first week"),
    ]),
    ("admin_users", "Admin: Users & Data", "👤", "Correcting, resetting, importing, exporting", [
        ("/setbaseline user: starting_xp:", "Set/correct someone's all-time starting XP"),
        ("/fullreset user: [clear_history]", "Reset both weekly and all-time totals for one user"),
        ("/removeuser user:<@user>", "Wipe one user's tracking data"),
        ("/removeuserbyid discord_id:<id>", "Same, but by raw ID — works for departed members"),
        ("/removeallusers", "Wipe everyone's tracking data (confirmation required)"),
        ("/removestaleusers", "Bulk-remove everyone no longer in the server"),
        ("/importxp file: [existing_users] [announce_milestones]", "Bulk-set XP from a CSV/Excel file"),
        ("/exportdata [file_format]", "Download everyone's stats as CSV/Excel"),
        ("/undoimport", "Revert the entire last import in one action"),
    ]),
    ("admin_roles", "Admin: XP Roles", "🏅", "Auto-assigning roles at XP milestones", [
        ("/addxprole xp: role: [exclusive]", "Auto-assign a role at an XP threshold"),
        ("/removexprole role:<@role>", "Stop auto-assigning a role"),
        ("/syncxproles", "Re-check everyone's roles right now"),
    ]),
    ("admin_display", "Admin: Display", "🎨", "How leaderboards and status look", [
        ("/togglerecentrate enabled:", "Show/hide the \"Recent Rate\" field"),
        ("/togglecompactleaderboard enabled:", "Short vs. full-detail leaderboard entries"),
        ("/toggleleaderboardpagination enabled:", "Previous/Next paging vs. a single truncated embed"),
    ]),
    ("admin_automation", "Admin: Channels & Automation", "📢", "Restrictions, announcements, schedules, pings", [
        ("/setchannel channel:<#channel>", "Restrict tracking commands to one channel"),
        ("/clearchannel", "Remove the channel restriction"),
        ("/setannouncechannel channel:<#channel>", "Public posts when someone earns a role"),
        ("/clearannouncechannel", "Turn off milestone announcements"),
        ("/setweeklypost channel:<#channel>", "Auto-post the leaderboard before each week ends"),
        ("/clearweeklypost", "Turn off the auto-posted leaderboard"),
        ("/setinactivitychannel channel:<#channel>", "Ping people who haven't checked in"),
        ("/clearinactivitychannel", "Turn off inactivity pings"),
        ("/setinactivitythreshold days: hours: minutes:", "How early the inactivity ping fires"),
        ("/toggleinactivitybehindpace enabled:", "Also flag people behind pace, not just zero-checkins"),
    ]),
]


def build_help_overview_embed() -> discord.Embed:
    embed = discord.Embed(
        title="📖 XP Tracker — Help",
        description=(
            "This bot tracks Roblox XP manually, since it can't read the game directly — "
            "everyone self-reports with `/checkin`.\n\nUse the dropdown below to browse commands by category."
        ),
        color=discord.Color.blurple(),
    )
    for value, label, emoji, desc, commands in HELP_CATEGORIES:
        embed.add_field(name=f"{emoji} {label}", value=desc, inline=True)
    embed.set_footer(text="Most commands work for everyone; admin-only ones need Manage Server.")
    return embed


def build_help_category_embed(category: str) -> discord.Embed:
    for value, label, emoji, desc, commands in HELP_CATEGORIES:
        if value == category:
            lines = [f"`{cmd}`\n{d}" for cmd, d in commands]
            embed = discord.Embed(
                title=f"{emoji} {label}",
                description=desc + "\n\n" + "\n\n".join(lines),
                color=discord.Color.blurple(),
            )
            return embed
    return build_help_overview_embed()


async def go_help_home(interaction: discord.Interaction):
    embed = build_help_overview_embed()
    view = HelpView()
    await interaction.response.edit_message(embed=embed, view=view)


class HelpCategorySelect(discord.ui.Select):
    def __init__(self):
        options = [
            discord.SelectOption(label=label, value=value, emoji=emoji, description=desc)
            for value, label, emoji, desc, commands in HELP_CATEGORIES
        ]
        super().__init__(placeholder="Choose a category to see its commands...", options=options, min_values=1, max_values=1)

    async def callback(self, interaction: discord.Interaction):
        embed = build_help_category_embed(self.values[0])
        view = HelpCategoryView()
        await interaction.response.edit_message(embed=embed, view=view)


class HelpView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=300)
        self.add_item(HelpCategorySelect())


class HelpCategoryView(discord.ui.View):
    def __init__(self):
        super().__init__(timeout=300)
        self.add_item(HelpCategorySelect())

    @discord.ui.button(label="⬅ Overview", style=discord.ButtonStyle.secondary, row=1)
    async def back(self, interaction: discord.Interaction, button: discord.ui.Button):
        await go_help_home(interaction)


SETTINGS_CATEGORIES = [
    ("requirement_week", "Requirement & Week", "⏱️", "Weekly XP goal, syncing everyone's week, proration"),
    ("display", "Display", "🎨", "Recent rate, compact leaderboard, pagination"),
    ("roles", "XP Roles", "🏅", "Milestone roles tied to XP thresholds"),
    ("automation", "Channels & Automation", "📢", "Tracking channel, announcements, schedules, pings"),
    ("danger", "Danger Zone", "⚠️", "Bulk-remove tracked users"),
]


async def go_settings_home(interaction: discord.Interaction, guild_id: int):
    data = load_data()
    guild_data = get_guild(data, guild_id)
    embed = build_settings_embed(guild_data)
    view = SettingsHomeView(guild_id)
    await interaction.response.edit_message(embed=embed, view=view)


class RequirementWeekView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=300)
        self.guild_id = guild_id

    @discord.ui.button(label="⬅ Back", style=discord.ButtonStyle.secondary)
    async def back(self, interaction: discord.Interaction, button: discord.ui.Button):
        await go_settings_home(interaction, self.guild_id)

    @discord.ui.button(label="Set Requirement", style=discord.ButtonStyle.primary, emoji="✏️")
    async def set_requirement(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(RequirementModal(self.guild_id))

    @discord.ui.button(label="Sync Week (Everyone)", style=discord.ButtonStyle.primary, emoji="🕒")
    async def sync_week(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(SyncWeekModal(self.guild_id))

    @discord.ui.button(label="Prorate Threshold", style=discord.ButtonStyle.primary, emoji="⏳")
    async def set_prorate_threshold(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.send_modal(ProrateThresholdModal(self.guild_id))


class DisplaySettingsView(discord.ui.View):
    def __init__(self, guild_id: int, show_recent_rate: bool, compact_leaderboard: bool, leaderboard_pagination: bool):
        super().__init__(timeout=300)
        self.guild_id = guild_id
        self._set_recent_rate_label(show_recent_rate)
        self._set_compact_label(compact_leaderboard)
        self._set_pagination_label(leaderboard_pagination)

    def _set_recent_rate_label(self, enabled: bool):
        self.toggle_recent_rate.label = "Recent Rate: ON" if enabled else "Recent Rate: OFF"
        self.toggle_recent_rate.style = discord.ButtonStyle.success if enabled else discord.ButtonStyle.secondary

    def _set_compact_label(self, enabled: bool):
        self.toggle_compact_leaderboard.label = "Compact Leaderboard: ON" if enabled else "Compact Leaderboard: OFF"
        self.toggle_compact_leaderboard.style = discord.ButtonStyle.success if enabled else discord.ButtonStyle.secondary

    def _set_pagination_label(self, enabled: bool):
        self.toggle_pagination.label = "Pagination: ON" if enabled else "Pagination: OFF"
        self.toggle_pagination.style = discord.ButtonStyle.success if enabled else discord.ButtonStyle.secondary

    @discord.ui.button(label="⬅ Back", style=discord.ButtonStyle.secondary)
    async def back(self, interaction: discord.Interaction, button: discord.ui.Button):
        await go_settings_home(interaction, self.guild_id)

    @discord.ui.button(label="Recent Rate: ON", style=discord.ButtonStyle.success, emoji="⏱️")
    async def toggle_recent_rate(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        new_state = not show_recent_rate_enabled(guild_data)
        guild_data["show_recent_rate"] = new_state
        save_data(data)
        self._set_recent_rate_label(new_state)
        embed = build_settings_embed(guild_data)
        await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Compact Leaderboard: OFF", style=discord.ButtonStyle.secondary, emoji="📏")
    async def toggle_compact_leaderboard(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        new_state = not compact_leaderboard_enabled(guild_data)
        guild_data["compact_leaderboard"] = new_state
        save_data(data)
        self._set_compact_label(new_state)
        embed = build_settings_embed(guild_data)
        await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Pagination: OFF", style=discord.ButtonStyle.secondary, emoji="📖")
    async def toggle_pagination(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        new_state = not leaderboard_pagination_enabled(guild_data)
        guild_data["leaderboard_pagination"] = new_state
        save_data(data)
        self._set_pagination_label(new_state)
        embed = build_settings_embed(guild_data)
        await interaction.response.edit_message(embed=embed, view=self)


class RolesSettingsView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=300)
        self.guild_id = guild_id

    @discord.ui.button(label="⬅ Back", style=discord.ButtonStyle.secondary)
    async def back(self, interaction: discord.Interaction, button: discord.ui.Button):
        await go_settings_home(interaction, self.guild_id)

    @discord.ui.button(label="Add XP Role", style=discord.ButtonStyle.primary, emoji="🏅")
    async def add_xp_role(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = AddXPRoleSelectView(self.guild_id)
        await interaction.response.send_message("Select a role to assign at an XP threshold:", view=view, ephemeral=True)

    @discord.ui.button(label="Remove XP Role", style=discord.ButtonStyle.secondary, emoji="🏅")
    async def remove_xp_role(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = RemoveXPRoleSelectView(self.guild_id)
        await interaction.response.send_message("Select a role to stop auto-assigning:", view=view, ephemeral=True)

    @discord.ui.button(label="Sync XP Roles Now", style=discord.ButtonStyle.secondary, emoji="🔁")
    async def sync_xp_roles(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        if not get_xp_roles(guild_data):
            await interaction.response.send_message("No XP milestone roles are configured yet.", ephemeral=True)
            return
        await interaction.response.defer(thinking=True, ephemeral=True)
        checked = added_count = removed_count = 0
        for uid, entry in guild_data["users"].items():
            member = interaction.guild.get_member(int(uid))
            if not member:
                continue
            checked += 1
            result = await apply_xp_roles(interaction.guild, member, guild_data, entry["current_xp"], announce=False)
            added_count += len(result["added"])
            removed_count += len(result["removed"])
        save_data(data)
        await interaction.followup.send(
            f"Re-checked **{checked}** member(s). Added **{added_count}**, removed **{removed_count}**.",
            ephemeral=True,
        )


class AutomationSettingsView(discord.ui.View):
    def __init__(self, guild_id: int, inactivity_behind_pace: bool):
        super().__init__(timeout=300)
        self.guild_id = guild_id
        self._set_behind_pace_label(inactivity_behind_pace)

    def _set_behind_pace_label(self, enabled: bool):
        self.toggle_behind_pace.label = "Inactivity: +Behind Pace" if enabled else "Inactivity: Zero Checkins Only"
        self.toggle_behind_pace.style = discord.ButtonStyle.success if enabled else discord.ButtonStyle.secondary

    @discord.ui.button(label="⬅ Back", style=discord.ButtonStyle.secondary, row=0)
    async def back(self, interaction: discord.Interaction, button: discord.ui.Button):
        await go_settings_home(interaction, self.guild_id)

    @discord.ui.button(label="Tracking Channel", style=discord.ButtonStyle.primary, emoji="📌", row=0)
    async def set_tracking_channel(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = SetTrackingChannelView(self.guild_id)
        await interaction.response.send_message("Select a channel:", view=view, ephemeral=True)

    @discord.ui.button(label="Announce Channel", style=discord.ButtonStyle.primary, emoji="📢", row=0)
    async def set_announce_channel(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = SetAnnounceChannelView(self.guild_id)
        await interaction.response.send_message("Select a channel:", view=view, ephemeral=True)

    @discord.ui.button(label="Weekly Post Channel", style=discord.ButtonStyle.primary, emoji="🗓️", row=1)
    async def set_weekly_post_channel(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = SetWeeklyPostChannelView(self.guild_id)
        await interaction.response.send_message("Select a channel:", view=view, ephemeral=True)

    @discord.ui.button(label="Inactivity Channel", style=discord.ButtonStyle.primary, emoji="⏰", row=1)
    async def set_inactivity_channel(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = SetInactivityChannelView(self.guild_id)
        await interaction.response.send_message("Select a channel:", view=view, ephemeral=True)

    @discord.ui.button(label="Admin Log Channel", style=discord.ButtonStyle.primary, emoji="🛡️", row=1)
    async def set_admin_log_channel(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = SetAdminLogChannelView(self.guild_id)
        await interaction.response.send_message("Select a channel:", view=view, ephemeral=True)

    @discord.ui.button(label="Inactivity: Zero Checkins Only", style=discord.ButtonStyle.secondary, emoji="🎯", row=2)
    async def toggle_behind_pace(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        new_state = not inactivity_includes_behind_pace(guild_data)
        guild_data["inactivity_include_behind_pace"] = new_state
        save_data(data)
        self._set_behind_pace_label(new_state)
        embed = build_settings_embed(guild_data)
        await interaction.response.edit_message(embed=embed, view=self)


class DangerZoneSettingsView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=300)
        self.guild_id = guild_id

    @discord.ui.button(label="⬅ Back", style=discord.ButtonStyle.secondary)
    async def back(self, interaction: discord.Interaction, button: discord.ui.Button):
        await go_settings_home(interaction, self.guild_id)

    @discord.ui.button(label="Remove All Users", style=discord.ButtonStyle.danger, emoji="🗑️")
    async def remove_all_users(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        count = len(guild_data["users"])
        if count == 0:
            await interaction.response.send_message("No one is being tracked yet — nothing to remove.", ephemeral=True)
            return
        view = ConfirmRemoveAllView(self.guild_id, count)
        await interaction.response.send_message(
            f"⚠️ This will permanently delete tracking data for **all {count} tracked user(s)** "
            f"in this server. This can't be undone. Continue?",
            view=view,
            ephemeral=True,
        )
        view.message = await interaction.original_response()

    @discord.ui.button(label="Remove Departed Members", style=discord.ButtonStyle.danger, emoji="👋")
    async def remove_stale_users(self, interaction: discord.Interaction, button: discord.ui.Button):
        await interaction.response.defer(thinking=True, ephemeral=True)
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        if not interaction.guild.chunked:
            try:
                await interaction.guild.chunk()
            except discord.ClientException:
                pass
        stale_uids = [uid for uid in guild_data["users"] if interaction.guild.get_member(int(uid)) is None]
        if not stale_uids:
            await interaction.followup.send("No stale entries found — everyone tracked is still in the server.", ephemeral=True)
            return
        view = ConfirmRemoveStaleView(self.guild_id, stale_uids)
        await interaction.followup.send(
            f"⚠️ Found **{len(stale_uids)}** tracked user(s) no longer in this server. Continue?",
            view=view,
            ephemeral=True,
        )
        view.message = await interaction.original_response()


def build_settings_category_view(category: str, guild_id: int, guild_data: dict) -> discord.ui.View:
    if category == "requirement_week":
        return RequirementWeekView(guild_id)
    if category == "display":
        return DisplaySettingsView(
            guild_id,
            show_recent_rate_enabled(guild_data),
            compact_leaderboard_enabled(guild_data),
            leaderboard_pagination_enabled(guild_data),
        )
    if category == "roles":
        return RolesSettingsView(guild_id)
    if category == "automation":
        return AutomationSettingsView(guild_id, inactivity_includes_behind_pace(guild_data))
    if category == "danger":
        return DangerZoneSettingsView(guild_id)
    return SettingsHomeView(guild_id)


class SettingsCategorySelect(discord.ui.Select):
    def __init__(self):
        options = [
            discord.SelectOption(label=label, value=value, emoji=emoji, description=desc)
            for value, label, emoji, desc in SETTINGS_CATEGORIES
        ]
        super().__init__(placeholder="Choose a settings category...", options=options, min_values=1, max_values=1)

    async def callback(self, interaction: discord.Interaction):
        guild_id = self.view.guild_id
        data = load_data()
        guild_data = get_guild(data, guild_id)
        view = build_settings_category_view(self.values[0], guild_id, guild_data)
        embed = build_settings_embed(guild_data)
        await interaction.response.edit_message(embed=embed, view=view)


class SettingsHomeView(discord.ui.View):
    def __init__(self, guild_id: int):
        super().__init__(timeout=300)
        self.guild_id = guild_id
        self.add_item(SettingsCategorySelect())

    @discord.ui.button(label="Manage a User", style=discord.ButtonStyle.secondary, emoji="👤")
    async def manage_user(self, interaction: discord.Interaction, button: discord.ui.Button):
        view = UserSelectView(self.guild_id)
        await interaction.response.send_message("Select a user:", view=view, ephemeral=True)

    @discord.ui.button(label="Refresh", style=discord.ButtonStyle.secondary, emoji="🔄")
    async def refresh(self, interaction: discord.Interaction, button: discord.ui.Button):
        data = load_data()
        guild_data = get_guild(data, self.guild_id)
        embed = build_settings_embed(guild_data)
        await interaction.response.edit_message(embed=embed, view=self)

# ---------------------------------------------------------------------------
# Bot setup
# ---------------------------------------------------------------------------

intents = discord.Intents.default()
intents.members = True  # required so guild.members includes everyone, not just a cached subset — needed for /importxp username matching
bot = commands.Bot(command_prefix="!", intents=intents)

GUILD_ID = os.environ.get("GUILD_ID")  # optional, for instant command sync during testing


SCHEDULER_INTERVAL_MINUTES = 15
# How far ahead of the exact week-end moment we're willing to fire the
# auto-post — wider than the loop interval so timing jitter can't cause us
# to skip the window entirely.
WEEKLY_POST_WINDOW = timedelta(minutes=SCHEDULER_INTERVAL_MINUTES + 5)


@tasks.loop(minutes=SCHEDULER_INTERVAL_MINUTES)
async def scheduled_checks():
    """Runs periodically for every guild the bot is in, checking two
    independent things — both only apply to guilds with a shared week
    (set via /setweekprogressall), since neither has a single meaningful
    trigger moment without one:

    1. Scheduled weekly leaderboard post — fires once, shortly BEFORE the
       week actually rolls over, so it captures the week's final results
       rather than the just-reset (near-zero) numbers of the new week.
    2. Inactivity ping — fires once per week when the configured amount
       of time remains before the week ends, listing anyone with zero
       checkins so far that week.

    Each fires at most once per week per guild, tracked via a marker
    storing which week's boundary it already handled."""
    data = load_data()
    changed = False

    for guild in bot.guilds:
        guild_data = get_guild(data, guild.id)
        anchor = guild_data.get("week_anchor")
        if not anchor:
            continue  # both features need a shared week to have one clear trigger moment

        week_start = parse_iso(anchor)
        now = utcnow()
        while now - week_start >= timedelta(days=7):
            week_start += timedelta(days=7)
        week_end = week_start + timedelta(days=7)
        time_until_end = week_end - now
        week_marker = iso(week_start)

        # --- Scheduled weekly post ---
        post_channel_id = get_weekly_post_channel_id(guild_data)
        if (
            post_channel_id
            and timedelta(0) < time_until_end <= WEEKLY_POST_WINDOW
            and guild_data.get("last_weekly_post_marker") != week_marker
        ):
            channel = guild.get_channel(post_channel_id)
            if channel:
                lines, extra_field, footer = build_weekly_leaderboard_lines(guild, guild_data)
                embed, view = build_leaderboard_payload(
                    guild_data, "📊 Weekly XP Leaderboard — Week Wrap-Up", lines, discord.Color.gold(), footer, extra_field
                )
                if embed:
                    try:
                        await channel.send(embed=embed, view=view)
                    except discord.HTTPException:
                        pass
            guild_data["last_weekly_post_marker"] = week_marker
            changed = True

        # --- Inactivity ping ---
        inactivity_channel_id = get_inactivity_channel_id(guild_data)
        if inactivity_channel_id:
            threshold = timedelta(hours=get_inactivity_threshold_hours(guild_data))
            include_behind_pace = inactivity_includes_behind_pace(guild_data)
            if (
                timedelta(0) < time_until_end <= threshold
                and guild_data.get("last_inactivity_ping_marker") != week_marker
            ):
                no_checkin_members = []
                behind_pace_members = []
                for uid, entry in guild_data["users"].items():
                    stats = compute_stats(entry, guild_data["requirement"], get_prorate_threshold_hours(guild_data))
                    member = guild.get_member(int(uid))
                    if not member:
                        continue
                    if stats["checkins_this_week"] == 0:
                        no_checkin_members.append(member)
                    elif include_behind_pace and not stats["on_track"]:
                        behind_pace_members.append(member)

                if no_checkin_members or behind_pace_members:
                    channel = guild.get_channel(inactivity_channel_id)
                    if channel:
                        time_left_str = format_duration(time_until_end.total_seconds() / 3600)
                        parts = [f"⏰ **Inactivity reminder** — the week ends in **{time_left_str}**."]
                        if no_checkin_members:
                            mentions = " ".join(m.mention for m in no_checkin_members)
                            parts.append(f"\n**Haven't checked in at all this week:**\n{mentions}")
                        if behind_pace_members:
                            mentions = " ".join(m.mention for m in behind_pace_members)
                            parts.append(f"\n**Checked in, but behind pace to hit the requirement:**\n{mentions}")
                        parts.append("\nRun `/checkin` before the week ends!")
                        try:
                            await channel.send("\n".join(parts))
                        except discord.HTTPException:
                            pass
                guild_data["last_inactivity_ping_marker"] = week_marker
                changed = True

    if changed:
        save_data(data)


@scheduled_checks.before_loop
async def before_scheduled_checks():
    await bot.wait_until_ready()


PRESENCE_ROTATION_SECONDS = 45
_presence_index = 0


def build_presence_statuses() -> list:
    """Builds the rotation of (ActivityType, text) pairs shown under the
    bot's name. Recomputed each rotation so counts stay current — cheap
    since it's just reading the already-loaded data file, not hitting
    Discord's API. A bot can only show simple text next to an activity
    verb (Playing/Watching/Listening/Competing) — the image+buttons style
    of Rich Presence is a feature for user game clients, not available to
    bots via the API."""
    data = load_data()
    total_tracked = sum(len(get_guild(data, g.id)["users"]) for g in bot.guilds)
    return [
        (discord.ActivityType.watching, f"{total_tracked} XP trackers"),
        (discord.ActivityType.listening, "/checkin"),
        (discord.ActivityType.playing, f"{len(bot.guilds)} servers"),
        (discord.ActivityType.watching, "/settings for admin tools"),
    ]


@tasks.loop(seconds=PRESENCE_ROTATION_SECONDS)
async def rotate_presence():
    global _presence_index
    statuses = build_presence_statuses()
    if not statuses:
        return
    activity_type, text = statuses[_presence_index % len(statuses)]
    _presence_index += 1
    try:
        await bot.change_presence(activity=discord.Activity(type=activity_type, name=text))
    except discord.HTTPException:
        pass


@rotate_presence.before_loop
async def before_rotate_presence():
    await bot.wait_until_ready()


@bot.event
async def on_ready():
    try:
        if GUILD_ID:
            guild_obj = discord.Object(id=int(GUILD_ID))
            bot.tree.copy_global_to(guild=guild_obj)
            synced = await bot.tree.sync(guild=guild_obj)
        else:
            synced = await bot.tree.sync()
        print(f"Synced {len(synced)} commands. Logged in as {bot.user}.")
    except Exception as e:
        print(f"Sync error: {e}")

    if not scheduled_checks.is_running():
        scheduled_checks.start()
    if not rotate_presence.is_running():
        rotate_presence.start()


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------

@bot.tree.command(name="help", description="Show all bot commands, organized by category")
async def help_cmd(interaction: discord.Interaction):
    # Deliberately NOT gated by @require_tracking_channel() — if someone's
    # in the wrong channel, /help is exactly what should still work
    # everywhere so they can find the right one.
    embed = build_help_overview_embed()
    view = HelpView()
    await interaction.response.send_message(embed=embed, view=view, ephemeral=True)


@bot.tree.command(name="checkin", description="Record current total Roblox XP — yours, or (admin) someone else's")
@app_commands.describe(
    xp="Current total XP in-game right now",
    user="Check in on behalf of someone else — admin only, defaults to you",
)
@require_tracking_channel()
async def checkin(interaction: discord.Interaction, xp: int, user: Optional[discord.Member] = None):
    if xp < 0:
        await interaction.response.send_message("XP can't be negative.", ephemeral=True)
        return

    target = user or interaction.user
    if user is not None and user.id != interaction.user.id:
        if not interaction.user.guild_permissions.manage_guild:
            await interaction.response.send_message(
                "You need the **Manage Server** permission to check in on behalf of someone else.", ephemeral=True
            )
            return

    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    uid = str(target.id)
    entry = guild_data["users"].get(uid)

    admin_note = f" (checked in by {interaction.user.display_name})" if target.id != interaction.user.id else ""

    if entry is None:
        entry = create_user(guild_data, target.id, xp)
        save_data(data)
        role_result = await apply_xp_roles(interaction.guild, target, guild_data, xp)
        role_note = (
            f"\n🎉 Reached: {', '.join(r.mention for r in role_result['added'])}"
            if role_result["added"] else ""
        )
        embed = discord.Embed(
            title="🟢 Tracking Started",
            description=f"Recorded starting XP of **{xp:,}** for {target.display_name}{admin_note}. "
                        f"Check in again later and I'll calculate {'their' if admin_note else 'your'} gain and pace."
                        f"{role_note}",
            color=discord.Color.blue(),
        )
        await interaction.response.send_message(embed=embed)
        return

    if xp < entry["current_xp"]:
        note = (f"⚠️ This is lower than the last recorded XP ({entry['current_xp']:,}). "
                f"Recorded anyway — use `/resetweek` (admin) if this was an XP reset in-game.")
    else:
        note = None

    last = record_checkin(entry, xp)
    stats = compute_stats(entry, guild_data["requirement"], get_prorate_threshold_hours(guild_data))
    save_data(data)
    role_result = await apply_xp_roles(interaction.guild, target, guild_data, xp)

    extra = None
    if last:
        gained = xp - last["xp"]
        elapsed = utcnow() - parse_iso(last["time"])
        hours = elapsed.total_seconds() / 3600
        duration_str = format_duration(hours)
        if hours >= 1 / 6:  # 10 minutes — same threshold used elsewhere, avoids wild extrapolated rates
            rate_str = f" ({gained/hours:,.1f} XP/hr)"
        else:
            rate_str = " (too soon for a reliable rate)"
        extra = ("Since Last Checkin", f"+{gained:,} XP over {duration_str}{rate_str}")

    embed = build_status_embed(
        target, entry, stats, guild_data["requirement"],
        extra_field=extra, show_recent_rate=show_recent_rate_enabled(guild_data),
    )
    if role_result["added"] or role_result["removed"]:
        parts = []
        if role_result["added"]:
            parts.append(f"Earned: {', '.join(r.mention for r in role_result['added'])}")
        if role_result["removed"]:
            parts.append(f"Replaced: {', '.join(r.mention for r in role_result['removed'])}")
        embed.add_field(name="🎉 Role Update", value="\n".join(parts), inline=False)
    if admin_note:
        embed.set_footer(text=f"Checked in by {interaction.user.display_name}")
    if note:
        embed.description = note
    await interaction.response.send_message(embed=embed)


@bot.tree.command(name="status", description="Check weekly XP progress for yourself or someone else")
@app_commands.describe(user="User to check (defaults to you)")
@require_tracking_channel()
async def status(interaction: discord.Interaction, user: Optional[discord.Member] = None):
    target = user or interaction.user
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    entry = guild_data["users"].get(str(target.id))

    if entry is None:
        await interaction.response.send_message(
            f"{target.display_name} hasn't checked in yet. Use `/checkin` to start tracking.",
            ephemeral=True,
        )
        return

    stats = compute_stats(entry, guild_data["requirement"], get_prorate_threshold_hours(guild_data))
    save_data(data)  # persist any lazy week rollover
    embed = build_status_embed(target, entry, stats, guild_data["requirement"], show_recent_rate=show_recent_rate_enabled(guild_data))
    await interaction.response.send_message(embed=embed)


def build_weekly_leaderboard_lines(guild: discord.Guild, guild_data: dict) -> tuple:
    """Returns (lines, extra_field, footer) for the weekly leaderboard.
    Shared between the /weeklyleaderboard command and the scheduled
    auto-post task. Caller is responsible for calling save_data() after,
    since building the lines runs ensure_current_week() on every entry."""
    compact = compact_leaderboard_enabled(guild_data)
    rows = []
    for uid, entry in guild_data["users"].items():
        stats = compute_stats(entry, guild_data["requirement"], get_prorate_threshold_hours(guild_data))
        rows.append((uid, entry, stats))

    rows.sort(key=lambda r: r[2]["gained_this_week"], reverse=True)
    lines = []
    for i, (uid, entry, stats) in enumerate(rows, start=1):
        member = guild.get_member(int(uid))
        name = member.display_name if member else f"<@{uid}>"
        icon = "✅" if stats["on_track"] else "⚠️"
        prorated = " 🆕" if stats["is_prorated"] else ""
        if compact:
            lines.append(f"{i}. **{name}** — {stats['gained_this_week']:,} XP {icon}{prorated}")
        else:
            low_data = " 🔸" if stats["checkins_this_week"] <= 1 else ""
            lines.append(
                f"**{i}. {name}** — {stats['gained_this_week']:,} XP {icon}{prorated}{low_data} "
                f"(proj. {stats['projected_total']:,.0f})"
            )

    anchor = guild_data.get("week_anchor")
    if anchor:
        week_start = parse_iso(anchor)
        now = utcnow()
        while now - week_start >= timedelta(days=7):  # roll forward to the current window, display-only
            week_start += timedelta(days=7)
        extra_field = ("📅 Week Runs", week_range_str(week_start))
    else:
        extra_field = (
            "📅 Week Runs",
            "Varies per person — no shared week set. Check `/status` for individual week ranges.",
        )

    footer = f"Weekly requirement: {guild_data['requirement']:,} XP"
    footer += "  •  🔸 = only 1 checkin this week" if not compact else ""
    footer += "  •  🆕 = prorated first week"
    return lines, extra_field, footer


def build_total_leaderboard_lines(guild: discord.Guild, guild_data: dict) -> list:
    """Returns lines for the all-time total leaderboard. Caller is
    responsible for save_data() after, since get_baseline() can lazily
    migrate older entries."""
    compact = compact_leaderboard_enabled(guild_data)
    rows = []
    for uid, entry in guild_data["users"].items():
        baseline_xp, baseline_time = get_baseline(entry)
        total_gained = entry["current_xp"] - baseline_xp
        tracked_days = max((utcnow() - parse_iso(baseline_time)).total_seconds() / 86400, 0)
        rows.append((uid, entry, baseline_xp, total_gained, tracked_days))

    rows.sort(key=lambda r: r[1]["current_xp"], reverse=True)  # rank by actual total XP, not just gain
    lines = []
    for i, (uid, entry, baseline_xp, total_gained, tracked_days) in enumerate(rows, start=1):
        member = guild.get_member(int(uid))
        name = member.display_name if member else f"<@{uid}>"
        if compact:
            lines.append(f"{i}. **{name}** — {entry['current_xp']:,} XP (+{total_gained:,})")
        else:
            avg_per_day = total_gained / tracked_days if tracked_days >= 1 / 24 else 0
            lines.append(
                f"**{i}. {name}** — {entry['current_xp']:,} XP total "
                f"(started at {baseline_xp:,}, +{total_gained:,} gained over {format_duration(tracked_days * 24)}, "
                f"avg {avg_per_day:,.0f}/day)"
            )
    return lines


@bot.tree.command(name="weeklyleaderboard", description="Show this week's XP leaderboard")
@require_tracking_channel()
async def weeklyleaderboard(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    lines, extra_field, footer = build_weekly_leaderboard_lines(interaction.guild, guild_data)
    save_data(data)

    embed, view = build_leaderboard_payload(
        guild_data, "📊 Weekly XP Leaderboard", lines, discord.Color.gold(), footer, extra_field
    )
    if embed is None:
        await interaction.response.send_message("No one is being tracked yet. Use `/checkin` to get started.")
        return
    if view:
        await interaction.response.send_message(embed=embed, view=view)
    else:
        await interaction.response.send_message(embed=embed)


@bot.tree.command(name="totalleaderboard", description="Show all-time total XP, including each person's starting point")
@require_tracking_channel()
async def totalleaderboard(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    lines = build_total_leaderboard_lines(interaction.guild, guild_data)
    save_data(data)  # persist any lazy baseline migration

    footer = "Ranked by actual total XP — use /setbaseline to correct a user's starting point."
    embed, view = build_leaderboard_payload(
        guild_data, "🏆 All-Time Total XP Leaderboard", lines, discord.Color.purple(), footer
    )
    if embed is None:
        await interaction.response.send_message("No one is being tracked yet. Use `/checkin` to get started.")
        return
    if view:
        await interaction.response.send_message(embed=embed, view=view)
    else:
        await interaction.response.send_message(embed=embed)


@bot.tree.command(name="crewtotals", description="Show the whole crew's combined XP — total and this week's gain")
@require_tracking_channel()
async def crewtotals(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)

    if not guild_data["users"]:
        await interaction.response.send_message("No one is being tracked yet. Use `/checkin` to get started.")
        return

    total_current_xp = 0
    total_gained_week = 0
    total_alltime_gained = 0
    on_track_count = 0
    member_count = len(guild_data["users"])

    for uid, entry in guild_data["users"].items():
        stats = compute_stats(entry, guild_data["requirement"], get_prorate_threshold_hours(guild_data))
        baseline_xp, _ = get_baseline(entry)
        total_current_xp += entry["current_xp"]
        total_gained_week += stats["gained_this_week"]
        total_alltime_gained += entry["current_xp"] - baseline_xp
        if stats["on_track"]:
            on_track_count += 1
    save_data(data)  # persist any lazy baseline migration or week rollovers triggered above

    avg_weekly = total_gained_week / member_count if member_count else 0

    embed = discord.Embed(title="👥 Crew XP Totals", color=discord.Color.blurple())
    embed.add_field(name="Total XP (crew)", value=f"{total_current_xp:,}", inline=True)
    embed.add_field(name="Gained This Week (crew)", value=f"{total_gained_week:,}", inline=True)
    embed.add_field(name="All-Time Gained (crew)", value=f"{total_alltime_gained:,}", inline=True)
    embed.add_field(name="Tracked Members", value=str(member_count), inline=True)
    embed.add_field(name="On Track This Week", value=f"{on_track_count}/{member_count}", inline=True)
    embed.add_field(name="Avg XP/Member This Week", value=f"{avg_weekly:,.0f}", inline=True)
    embed.set_footer(text="Includes members no longer in the server — their tracked XP still counts toward crew totals.")
    await interaction.response.send_message(embed=embed)


@bot.tree.command(name="history", description="Show recent XP checkins for a user")
@app_commands.describe(user="User to check (defaults to you)")
@require_tracking_channel()
async def history(interaction: discord.Interaction, user: Optional[discord.Member] = None):
    target = user or interaction.user
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    entry = guild_data["users"].get(str(target.id))

    if entry is None:
        await interaction.response.send_message("No data yet for that user.", ephemeral=True)
        return

    checkins = entry["checkins"][-10:]
    lines = []
    for c in reversed(checkins):
        t = parse_iso(c["time"])
        lines.append(f"<t:{int(t.timestamp())}:R> — {c['xp']:,} XP")

    embed = discord.Embed(
        title=f"📜 Recent Checkins — {target.display_name}",
        description="\n".join(lines) or "No checkins recorded.",
        color=discord.Color.blue(),
    )
    await interaction.response.send_message(embed=embed)


@bot.tree.command(name="progresschart", description="Show a chart of XP progress over time for yourself or someone else")
@app_commands.describe(user="User to chart (defaults to you)")
@require_tracking_channel()
async def progresschart(interaction: discord.Interaction, user: Optional[discord.Member] = None):
    target = user or interaction.user
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    entry = guild_data["users"].get(str(target.id))

    if entry is None:
        await interaction.response.send_message("No data yet for that user.", ephemeral=True)
        return

    if plt is None:
        await interaction.response.send_message(
            "Chart generation requires the `matplotlib` package, which isn't installed. "
            "Run `pip install matplotlib` and restart the bot.",
            ephemeral=True,
        )
        return

    if len(entry["checkins"]) < 2:
        await interaction.response.send_message(
            f"{target.display_name} needs at least 2 checkins before a chart is possible — "
            f"there's only {len(entry['checkins'])} so far.",
            ephemeral=True,
        )
        return

    await interaction.response.defer(thinking=True)
    chart_buf = build_progress_chart(entry, target.display_name)
    file = discord.File(fp=chart_buf, filename="progress.png")
    embed = discord.Embed(title=f"📈 XP Progress — {target.display_name}", color=discord.Color.blurple())
    embed.set_image(url="attachment://progress.png")
    await interaction.followup.send(embed=embed, file=file)


@bot.tree.command(name="undo", description="Remove the most recent checkin (yours, or someone else's if you're an admin)")
@app_commands.describe(user="Whose checkin to undo — defaults to you. Undoing someone else requires Manage Server.")
@require_tracking_channel()
async def undo(interaction: discord.Interaction, user: Optional[discord.Member] = None):
    target = user or interaction.user

    if user is not None and user.id != interaction.user.id:
        if not interaction.user.guild_permissions.manage_guild:
            await interaction.response.send_message(
                "You need the **Manage Server** permission to undo someone else's checkin.", ephemeral=True
            )
            return

    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    entry = guild_data["users"].get(str(target.id))

    if entry is None or not entry["checkins"]:
        await interaction.response.send_message("No checkins to undo.", ephemeral=True)
        return

    if len(entry["checkins"]) == 1:
        # Undoing the only checkin means wiping the user's tracking entirely.
        del guild_data["users"][str(target.id)]
        save_data(data)
        if target.id != interaction.user.id:
            await log_admin_action(interaction.guild, guild_data, interaction.user, "Undo (removed only checkin)", target=target.mention)
        await interaction.response.send_message(
            f"Removed {target.display_name}'s only checkin — tracking has been reset. "
            f"Use `/checkin` to start again."
        )
        return

    removed = entry["checkins"].pop()
    entry["current_xp"] = entry["checkins"][-1]["xp"]
    save_data(data)
    if target.id != interaction.user.id:
        await log_admin_action(
            interaction.guild, guild_data, interaction.user, "Undo checkin",
            target=target.mention, details=f"Reverted {removed['xp']:,} XP to {entry['current_xp']:,} XP.",
        )

    stats = compute_stats(entry, guild_data["requirement"], get_prorate_threshold_hours(guild_data))
    embed = build_status_embed(target, entry, stats, guild_data["requirement"], show_recent_rate=show_recent_rate_enabled(guild_data))
    removed_time = int(parse_iso(removed["time"]).timestamp())
    embed.description = (
        f"Undid checkin of **{removed['xp']:,} XP** recorded <t:{removed_time}:R>. "
        f"Current XP reverted to **{entry['current_xp']:,}**."
    )
    await interaction.response.send_message(embed=embed)


@bot.tree.command(name="settings", description="[Admin] Open an interactive panel to change server settings")
@app_commands.checks.has_permissions(manage_guild=True)
async def settings_cmd(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    save_data(data)  # persist guild creation if this is the first time
    embed = build_settings_embed(guild_data)
    view = SettingsHomeView(interaction.guild_id)
    await interaction.response.send_message(embed=embed, view=view)


@bot.tree.command(
    name="togglerecentrate",
    description="[Admin] Show or hide the 'Recent Rate' field (can look wrong after a week-progress edit)",
)
@app_commands.describe(enabled="True to show Recent Rate, false to hide it")
@app_commands.checks.has_permissions(manage_guild=True)
async def togglerecentrate(interaction: discord.Interaction, enabled: bool):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["show_recent_rate"] = enabled
    save_data(data)
    state = "shown" if enabled else "hidden"
    await interaction.response.send_message(f"Recent Rate will now be **{state}** on `/status` and `/checkin`.")


@bot.tree.command(
    name="setproratethreshold",
    description="[Admin] Set how late someone must join to get a prorated (reduced) weekly requirement",
)
@app_commands.describe(
    days="Hours-remaining threshold, days part (0-7)",
    hours="Plus hours (0-23)",
    minutes="Plus minutes (0-59)",
)
@app_commands.checks.has_permissions(manage_guild=True)
async def setproratethreshold(interaction: discord.Interaction, days: int = 0, hours: int = 0, minutes: int = 0):
    try:
        delta = timedelta(days=days, hours=hours, minutes=minutes)
        if delta < timedelta(0) or delta > timedelta(days=7):
            raise ValueError("Total time must be between 0 and 7 days.")
    except ValueError as e:
        await interaction.response.send_message(str(e), ephemeral=True)
        return

    threshold_hours = delta.total_seconds() / 3600
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["prorate_threshold_hours"] = threshold_hours
    save_data(data)

    if threshold_hours >= 168:
        note = "Any late join at all will now get a prorated requirement (this is the default)."
    elif threshold_hours <= 0:
        note = "Proration is now effectively **off** — everyone is held to the full flat requirement regardless of when they joined."
    else:
        note = (
            f"Someone must join with **{format_duration(threshold_hours)} or less** remaining in the "
            f"week for their requirement to be prorated. Joining earlier than that uses the full requirement."
        )
    await interaction.response.send_message(f"Prorate threshold set to {format_duration(threshold_hours)}. {note}")


@bot.tree.command(
    name="togglecompactleaderboard",
    description="[Admin] Use shorter leaderboard entries so the embed takes up less space",
)
@app_commands.describe(enabled="True for compact one-line entries, false for full detail")
@app_commands.checks.has_permissions(manage_guild=True)
async def togglecompactleaderboard(interaction: discord.Interaction, enabled: bool):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["compact_leaderboard"] = enabled
    save_data(data)
    state = "compact" if enabled else "full detail"
    await interaction.response.send_message(
        f"Leaderboards will now use **{state}** formatting on `/weeklyleaderboard` and `/totalleaderboard`."
    )


@bot.tree.command(name="setchannel", description="[Admin] Restrict tracking commands to a single channel")
@app_commands.describe(channel="Channel where /checkin, /status, and leaderboards will be allowed")
@app_commands.checks.has_permissions(manage_guild=True)
async def setchannel(interaction: discord.Interaction, channel: discord.TextChannel):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["restricted_channel_id"] = channel.id
    save_data(data)
    await interaction.response.send_message(
        f"Tracking commands (`/checkin`, `/status`, `/weeklyleaderboard`, `/totalleaderboard`, `/history`, "
        f"`/undo`) are now restricted to {channel.mention}. Admin/config commands still work anywhere."
    )


@bot.tree.command(name="clearchannel", description="[Admin] Remove the channel restriction — allow tracking commands anywhere")
@app_commands.checks.has_permissions(manage_guild=True)
async def clearchannel(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["restricted_channel_id"] = None
    save_data(data)
    await interaction.response.send_message("Channel restriction removed — tracking commands now work in any channel.")


@bot.tree.command(name="setannouncechannel", description="[Admin] Post a public message when someone earns an XP milestone role")
@app_commands.describe(channel="Channel where milestone announcements will be posted")
@app_commands.checks.has_permissions(manage_guild=True)
async def setannouncechannel(interaction: discord.Interaction, channel: discord.TextChannel):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["announce_channel_id"] = channel.id
    save_data(data)
    await interaction.response.send_message(
        f"Milestone role announcements will now be posted in {channel.mention}. "
        f"Bulk operations (imports, retroactive grants, syncs) stay quiet to avoid flooding the channel — "
        f"only organic checkins trigger a public announcement."
    )


@bot.tree.command(name="clearannouncechannel", description="[Admin] Turn off public milestone role announcements")
@app_commands.checks.has_permissions(manage_guild=True)
async def clearannouncechannel(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["announce_channel_id"] = None
    save_data(data)
    await interaction.response.send_message("Milestone role announcements are now off.")


@bot.tree.command(name="setweeklypost", description="[Admin] Auto-post the weekly leaderboard shortly before each week ends")
@app_commands.describe(channel="Channel where the weekly leaderboard will be auto-posted")
@app_commands.checks.has_permissions(manage_guild=True)
async def setweeklypost(interaction: discord.Interaction, channel: discord.TextChannel):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    if not guild_data.get("week_anchor"):
        await interaction.response.send_message(
            "This requires a shared server week first — run `/setweekprogressall` at least once "
            "so there's a single week boundary to schedule around.",
            ephemeral=True,
        )
        return
    guild_data["weekly_post_channel_id"] = channel.id
    save_data(data)
    await interaction.response.send_message(
        f"The weekly leaderboard will now auto-post in {channel.mention} shortly before each week ends "
        f"(capturing the week's final results, not the just-reset numbers of the new week)."
    )


@bot.tree.command(name="clearweeklypost", description="[Admin] Turn off the auto-posted weekly leaderboard")
@app_commands.checks.has_permissions(manage_guild=True)
async def clearweeklypost(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["weekly_post_channel_id"] = None
    save_data(data)
    await interaction.response.send_message("Weekly leaderboard auto-posting is now off.")


@bot.tree.command(name="setinactivitychannel", description="[Admin] Ping people who haven't checked in as the week nears its end")
@app_commands.describe(channel="Channel where inactivity reminders will be posted")
@app_commands.checks.has_permissions(manage_guild=True)
async def setinactivitychannel(interaction: discord.Interaction, channel: discord.TextChannel):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    if not guild_data.get("week_anchor"):
        await interaction.response.send_message(
            "This requires a shared server week first — run `/setweekprogressall` at least once "
            "so there's a single week boundary to schedule around.",
            ephemeral=True,
        )
        return
    guild_data["inactivity_channel_id"] = channel.id
    save_data(data)
    threshold = get_inactivity_threshold_hours(guild_data)
    await interaction.response.send_message(
        f"Inactivity reminders will now post in {channel.mention} once **{format_duration(threshold)}** remain "
        f"in the week, @mentioning anyone with zero checkins so far. Adjust the timing with `/setinactivitythreshold`."
    )


@bot.tree.command(name="clearinactivitychannel", description="[Admin] Turn off inactivity reminder pings")
@app_commands.checks.has_permissions(manage_guild=True)
async def clearinactivitychannel(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["inactivity_channel_id"] = None
    save_data(data)
    await interaction.response.send_message("Inactivity reminder pings are now off.")


@bot.tree.command(name="setadminlogchannel", description="[Admin] Log destructive/data-altering admin actions to a channel")
@app_commands.describe(channel="Channel where admin action logs will be posted")
@app_commands.checks.has_permissions(manage_guild=True)
async def setadminlogchannel(interaction: discord.Interaction, channel: discord.TextChannel):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["admin_log_channel_id"] = channel.id
    save_data(data)
    await interaction.response.send_message(
        f"Admin actions (resets, removals, imports, requirement changes, etc.) will now be logged in {channel.mention}."
    )


@bot.tree.command(name="clearadminlogchannel", description="[Admin] Turn off the admin action log")
@app_commands.checks.has_permissions(manage_guild=True)
async def clearadminlogchannel(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["admin_log_channel_id"] = None
    save_data(data)
    await interaction.response.send_message("Admin action logging is now off.")


@bot.tree.command(name="setinactivitythreshold", description="[Admin] Set how long before week-end inactivity pings fire")
@app_commands.describe(
    days="Time-before-week-end threshold, days part (0-7)",
    hours="Plus hours (0-23)",
    minutes="Plus minutes (0-59)",
)
@app_commands.checks.has_permissions(manage_guild=True)
async def setinactivitythreshold(interaction: discord.Interaction, days: int = 0, hours: int = 0, minutes: int = 0):
    try:
        delta = parse_dhm_fields(str(days), str(hours), str(minutes))
    except ValueError as e:
        await interaction.response.send_message(str(e), ephemeral=True)
        return
    if delta <= timedelta(0):
        await interaction.response.send_message("Threshold must be greater than zero.", ephemeral=True)
        return

    threshold_hours = delta.total_seconds() / 3600
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["inactivity_threshold_hours"] = threshold_hours
    save_data(data)
    await interaction.response.send_message(
        f"Inactivity pings will now fire when **{format_duration(threshold_hours)}** remain before the week ends."
    )


@bot.tree.command(
    name="toggleinactivitybehindpace",
    description="[Admin] Also ping people who checked in but are behind pace, not just zero-checkin people",
)
@app_commands.describe(enabled="True to also include behind-pace members, false for zero-checkins only (default)")
@app_commands.checks.has_permissions(manage_guild=True)
async def toggleinactivitybehindpace(interaction: discord.Interaction, enabled: bool):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["inactivity_include_behind_pace"] = enabled
    save_data(data)
    if enabled:
        msg = (
            "Inactivity pings will now include **two groups**: people with zero checkins this week, "
            "and people who've checked in but are behind pace to hit the requirement."
        )
    else:
        msg = "Inactivity pings will only include people with **zero checkins** this week (default)."
    await interaction.response.send_message(msg)


@bot.tree.command(name="toggleleaderboardpagination", description="[Admin] Use Previous/Next buttons instead of truncating long leaderboards")
@app_commands.describe(enabled="True for paginated Previous/Next browsing, false for a single truncated embed")
@app_commands.checks.has_permissions(manage_guild=True)
async def toggleleaderboardpagination(interaction: discord.Interaction, enabled: bool):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    guild_data["leaderboard_pagination"] = enabled
    save_data(data)
    state = "paginated (Previous/Next buttons)" if enabled else "a single truncated embed"
    await interaction.response.send_message(f"Leaderboards will now use {state}.")


@bot.tree.command(name="setrequirement", description="[Admin] Set the weekly XP requirement for this server")
@app_commands.describe(xp="New weekly XP requirement")
@app_commands.checks.has_permissions(manage_guild=True)
async def setrequirement(interaction: discord.Interaction, xp: int):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    old_xp = guild_data["requirement"]
    guild_data["requirement"] = xp
    save_data(data)
    await log_admin_action(
        interaction.guild, guild_data, interaction.user, "Changed weekly requirement",
        details=f"{old_xp:,} XP → {xp:,} XP",
    )
    await interaction.response.send_message(f"Weekly requirement set to **{xp:,} XP**.")


@bot.tree.command(name="resetweek", description="[Admin] Manually reset a user's weekly tracking window")
@app_commands.describe(user="User whose weekly window should reset now")
@app_commands.checks.has_permissions(manage_guild=True)
async def resetweek(interaction: discord.Interaction, user: discord.Member):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    entry = guild_data["users"].get(str(user.id))

    if entry is None:
        await interaction.response.send_message("That user has no tracking data.", ephemeral=True)
        return

    entry["week_start_time"] = iso(utcnow())
    entry["week_start_xp"] = entry["current_xp"]
    save_data(data)
    await log_admin_action(interaction.guild, guild_data, interaction.user, "Reset weekly window", target=user.mention)
    await interaction.response.send_message(f"Weekly tracking window reset for {user.display_name}.")


@bot.tree.command(name="backup", description="[Admin] Download a full backup of ALL settings and tracked data for this server")
@app_commands.checks.has_permissions(manage_guild=True)
async def backup_cmd(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    save_data(data)  # persist any lazy guild creation

    payload = {
        "backup_version": 1,
        "backup_time": iso(utcnow()),
        "guild_id": interaction.guild_id,
        "guild_name": interaction.guild.name,
        "guild_data": guild_data,
    }
    json_bytes = io.BytesIO(json.dumps(payload, indent=2).encode("utf-8"))
    timestamp = utcnow().strftime("%Y%m%d_%H%M%S")
    filename = f"xp_backup_{timestamp}.json"

    await interaction.response.send_message(
        content=(
            f"Full backup — **{len(guild_data['users'])}** tracked user(s), plus every setting "
            f"(requirement, week sync, XP roles, channels, toggles). This is different from "
            f"`/exportdata`, which only exports current stats, not settings. Keep this file "
            f"somewhere safe; `/restore` can rebuild everything from it."
        ),
        file=discord.File(fp=json_bytes, filename=filename),
        ephemeral=True,
    )


@bot.tree.command(name="restore", description="[Admin] Restore ALL settings and tracked data from a /backup file — replaces everything")
@app_commands.describe(file="The .json file downloaded from /backup")
@app_commands.checks.has_permissions(manage_guild=True)
async def restore_cmd(interaction: discord.Interaction, file: discord.Attachment):
    await interaction.response.defer(thinking=True, ephemeral=True)

    try:
        raw = await file.read()
        payload = json.loads(raw.decode("utf-8"))
    except Exception:
        await interaction.followup.send(
            "Couldn't read that file — make sure it's a valid backup .json produced by `/backup`.", ephemeral=True
        )
        return

    if not isinstance(payload, dict) or "guild_data" not in payload:
        await interaction.followup.send(
            "That doesn't look like a valid backup file — missing the expected `guild_data` field.", ephemeral=True
        )
        return

    restored_guild_data = payload["guild_data"]
    if not isinstance(restored_guild_data, dict) or "users" not in restored_guild_data:
        await interaction.followup.send("That backup file's data looks malformed — can't safely restore it.", ephemeral=True)
        return

    user_count = len(restored_guild_data.get("users", {}))
    backup_time_str = payload.get("backup_time")
    try:
        backup_ts = int(parse_iso(backup_time_str).timestamp())
        time_str = f"<t:{backup_ts}:R>"
    except (TypeError, ValueError):
        time_str = "at an unknown time"

    view = ConfirmRestoreView(interaction.guild_id, restored_guild_data)
    await interaction.followup.send(
        f"⚠️ This will **completely replace** all settings and tracked data for this server with "
        f"a backup taken {time_str} (**{user_count}** tracked user(s) in that backup). Everything "
        f"currently in place — including anything that's happened since that backup — will be lost. "
        f"This can't be undone with `/undoimport` or any other command. Continue?",
        view=view,
    )
    view.message = await interaction.original_response()


@bot.tree.command(name="exportdata", description="[Admin] Export everyone's current stats to a CSV or Excel file")
@app_commands.describe(file_format="Output format — csv or xlsx")
@app_commands.choices(file_format=[
    app_commands.Choice(name="CSV", value="csv"),
    app_commands.Choice(name="Excel (.xlsx)", value="xlsx"),
])
@app_commands.checks.has_permissions(manage_guild=True)
async def exportdata(interaction: discord.Interaction, file_format: app_commands.Choice[str] = None):
    fmt = file_format.value if file_format else "csv"
    await interaction.response.defer(thinking=True)

    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)

    if not interaction.guild.chunked:
        try:
            await interaction.guild.chunk()
        except discord.ClientException:
            pass

    rows = []
    headers = [
        "discord_id", "username", "starting_xp", "baseline_xp", "total_gained",
        "week_start_xp", "gained_this_week", "last_checkin_time_utc",
    ]
    for uid, entry in guild_data["users"].items():
        member = interaction.guild.get_member(int(uid))
        username = member.name if member else ""
        baseline_xp, _ = get_baseline(entry)
        last_checkin = entry["checkins"][-1]["time"] if entry["checkins"] else ""
        rows.append([
            uid,
            username,
            entry["current_xp"],  # exported under the "starting_xp" header so /importxp can read it directly
            baseline_xp,
            entry["current_xp"] - baseline_xp,
            entry["week_start_xp"],
            entry["current_xp"] - entry["week_start_xp"],
            last_checkin,
        ])
    save_data(data)  # persist any lazy baseline migration triggered by get_baseline

    if not rows:
        await interaction.followup.send("No one is being tracked yet — nothing to export.", ephemeral=True)
        return

    timestamp = utcnow().strftime("%Y%m%d_%H%M%S")

    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerows(rows)
        file_bytes = io.BytesIO(buf.getvalue().encode("utf-8"))
        filename = f"xp_export_{timestamp}.csv"
    else:
        if openpyxl is None:
            await interaction.followup.send(
                "Excel export requires the `openpyxl` package, which isn't installed. "
                "Run `pip install openpyxl`, or use `file_format: CSV` instead.",
                ephemeral=True,
            )
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "XP Export"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        file_bytes = buf
        filename = f"xp_export_{timestamp}.xlsx"

    file_bytes.seek(0)
    await interaction.followup.send(
        content=f"Exported stats for **{len(rows)}** tracked user(s).",
        file=discord.File(fp=file_bytes, filename=filename),
    )


@bot.tree.command(
    name="importxp",
    description="[Admin] Bulk-set XP for many users from a CSV or Excel file",
)
@app_commands.describe(
    file="A .csv or .xlsx file. Needs a column for XP (starting_xp/xp) and either discord_id or username",
    existing_users="What to do for users already being tracked (new users are always registered either way)",
    announce_milestones="Post public milestone announcements for roles earned during this import (default: no, stay quiet)",
)
@app_commands.choices(existing_users=[
    app_commands.Choice(name="Skip them (default) — only register brand-new users", value="skip"),
    app_commands.Choice(name="Check in — update XP like a normal checkin, keeps their history", value="checkin"),
    app_commands.Choice(name="Reset — wipe and restart, like /fullreset", value="reset"),
])
@app_commands.checks.has_permissions(manage_guild=True)
async def importxp(
    interaction: discord.Interaction,
    file: discord.Attachment,
    existing_users: app_commands.Choice[str] = None,
    announce_milestones: bool = False,
):
    mode = existing_users.value if existing_users else "skip"
    await interaction.response.defer(thinking=True)

    guild = interaction.guild
    if not guild.chunked:
        try:
            await guild.chunk()
        except discord.ClientException:
            pass  # members intent not enabled — matching will fall back to whatever's cached

    try:
        raw = await file.read()
        rows = parse_spreadsheet_rows(file.filename, raw)
    except ValueError as e:
        await interaction.followup.send(str(e), ephemeral=True)
        return
    except Exception:
        await interaction.followup.send("Couldn't read that file — make sure it's a valid .csv or .xlsx.", ephemeral=True)
        return

    if not rows:
        await interaction.followup.send("That file has no data rows.", ephemeral=True)
        return

    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)

    # Snapshot everyone's state BEFORE this import touches anything, so the
    # whole batch can be reverted in one action via /undoimport — /undo only
    # ever handles one person's most recent checkin at a time, which isn't
    # enough to cleanly reverse a bulk import (especially "reset" mode,
    # which replaces checkin history outright rather than appending to it).
    guild_data["last_import_snapshot"] = {
        "time": iso(utcnow()),
        "users": copy.deepcopy(guild_data["users"]),
    }

    created, checked_in, reset_list, skipped, unmatched, ambiguous, bad_rows = [], [], [], [], [], [], []
    roles_assigned_count = 0

    for row in rows:
        xp_val = find_xp_for_row(row)
        if xp_val is None:
            bad_rows.append(row)
            continue

        label = row.get("discord_id") or row.get("username") or row.get("name") or "(unknown row)"
        member, match_type = find_member_for_row(guild, row)

        if member is None:
            if match_type.startswith("ambiguous"):
                ambiguous.append(str(label))
            else:
                unmatched.append(str(label))
            continue

        uid = str(member.id)
        existing = guild_data["users"].get(uid)
        tag = " (via nickname match)" if match_type == "substring" else ""

        if existing and mode == "skip":
            skipped.append(member.display_name)
            continue

        if existing and mode == "checkin":
            # Same as a normal /checkin: appends to their history and moves
            # current_xp, but leaves baseline_xp and week_start_xp alone —
            # this is what makes it a mass check-in rather than a reset.
            record_checkin(existing, xp_val)
            checked_in.append(f"{member.display_name} ({xp_val:,}){tag}")
        elif existing and mode == "reset":
            now = utcnow()
            existing["current_xp"] = xp_val
            existing["baseline_xp"] = xp_val
            existing["baseline_time"] = iso(now)
            existing["week_start_time"] = iso(now)
            existing["week_start_xp"] = xp_val
            existing["checkins"] = [{"time": iso(now), "xp": xp_val}]
            reset_list.append(f"{member.display_name} ({xp_val:,}){tag}")
        else:
            create_user(guild_data, member.id, xp_val)
            created.append(f"{member.display_name} ({xp_val:,}){tag}")

        role_result = await apply_xp_roles(guild, member, guild_data, xp_val, announce=announce_milestones)
        roles_assigned_count += len(role_result["added"])

    save_data(data)

    await log_admin_action(
        guild, guild_data, interaction.user, "Bulk import (/importxp)",
        details=(
            f"Mode: {mode}. Created: {len(created)}, Checked in: {len(checked_in)}, "
            f"Reset: {len(reset_list)}, Skipped: {len(skipped)}, Unmatched: {len(unmatched)}."
        ),
    )

    def fmt_list(items, limit=15):
        if not items:
            return "—"
        shown = "\n".join(items[:limit])
        if len(items) > limit:
            shown += f"\n…and {len(items) - limit} more"
        return shown

    embed = discord.Embed(title="📥 XP Import Results", color=discord.Color.green())
    embed.add_field(name=f"✅ Newly tracked ({len(created)})", value=fmt_list(created), inline=False)
    if mode == "checkin":
        embed.add_field(name=f"🔁 Checked in ({len(checked_in)})", value=fmt_list(checked_in), inline=False)
    elif mode == "reset":
        embed.add_field(name=f"♻️ Reset ({len(reset_list)})", value=fmt_list(reset_list), inline=False)
    else:
        embed.add_field(name=f"⏭️ Skipped — already tracked ({len(skipped)})", value=fmt_list(skipped), inline=False)
    if ambiguous:
        embed.add_field(
            name=f"⚠️ Ambiguous — multiple matches ({len(ambiguous)})",
            value=fmt_list(ambiguous) + "\nMatched more than one member — use a `discord_id` column for these.",
            inline=False,
        )
    embed.add_field(name=f"❓ Unmatched ({len(unmatched)})", value=fmt_list(unmatched), inline=False)
    if bad_rows:
        embed.add_field(
            name=f"⚠️ Rows with missing/invalid XP ({len(bad_rows)})",
            value="Make sure your XP column is named `starting_xp` or `xp`.",
            inline=False,
        )
    if roles_assigned_count:
        announce_note = (
            " Public announcements were posted for these."
            if announce_milestones and get_announce_channel_id(guild_data)
            else " Announcements stayed quiet for this import."
        )
        embed.add_field(
            name="🎉 Milestone Roles Assigned",
            value=f"{roles_assigned_count} role(s) granted based on configured XP thresholds.{announce_note}",
            inline=False,
        )
    embed.set_footer(text="Unmatched? Check the username matches their nickname/username, or add a discord_id column. Made a mistake? Run /undoimport to revert this entire batch.")
    await interaction.followup.send(embed=embed)


@bot.tree.command(name="undoimport", description="[Admin] Revert everyone's data back to how it looked before the last /importxp")
@app_commands.checks.has_permissions(manage_guild=True)
async def undoimport(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    snapshot = guild_data.get("last_import_snapshot")

    if not snapshot:
        await interaction.response.send_message(
            "No import to undo — either nothing's been imported yet, or the last import was already undone.",
            ephemeral=True,
        )
        return

    before_count = len(snapshot["users"])
    after_count = len(guild_data["users"])
    snapshot_time = int(parse_iso(snapshot["time"]).timestamp())

    view = ConfirmUndoImportView(interaction.guild_id)
    await interaction.response.send_message(
        f"⚠️ This will revert **everyone's** tracking data back to how it looked "
        f"<t:{snapshot_time}:R>, right before that import ran (currently **{after_count}** "
        f"tracked, will become **{before_count}**). Anything anyone did *after* that import — "
        f"checkins, corrections, role changes — will also be lost, not just the import itself. "
        f"This can only undo the *one* most recent import. Continue?",
        view=view,
    )
    view.message = await interaction.original_response()


@bot.tree.command(
    name="addxprole",
    description="[Admin] Auto-assign a role once someone reaches an XP milestone",
)
@app_commands.describe(
    xp="XP total required to earn this role",
    role="Role to grant once that XP is reached",
    exclusive="True (default): removed when a higher tier is reached. False: permanent, never removed.",
)
@app_commands.checks.has_permissions(manage_guild=True)
async def addxprole(interaction: discord.Interaction, xp: int, role: discord.Role, exclusive: bool = True):
    if xp < 0:
        await interaction.response.send_message("XP must be a positive number.", ephemeral=True)
        return

    bot_member = interaction.guild.me
    if role >= bot_member.top_role:
        await interaction.response.send_message(
            f"⚠️ I can't assign {role.mention} — it's positioned at or above my own highest role in "
            f"Server Settings → Roles. Move my role above it, then try again.",
            ephemeral=True,
        )
        return
    if not bot_member.guild_permissions.manage_roles:
        await interaction.response.send_message(
            "⚠️ I don't have the **Manage Roles** permission in this server. Re-invite me with that "
            "permission, or enable it for my role in Server Settings → Roles.",
            ephemeral=True,
        )
        return

    await interaction.response.defer(thinking=True)

    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    xp_roles = guild_data.setdefault("xp_roles", [])
    xp_roles[:] = [r for r in xp_roles if r["role_id"] != role.id]  # replace any existing mapping for this role
    xp_roles.append({"xp": xp, "role_id": role.id, "exclusive": exclusive})
    xp_roles.sort(key=lambda r: r["xp"])

    # Retroactively grant to anyone who already qualifies, so people don't
    # have to check in again just to receive a role that already applies to them.
    granted = 0
    for uid, entry in guild_data["users"].items():
        if entry["current_xp"] >= xp:
            member = interaction.guild.get_member(int(uid))
            if member and role not in member.roles:
                result = await apply_xp_roles(interaction.guild, member, guild_data, entry["current_xp"], announce=False)
                if result["added"]:
                    granted += 1
    save_data(data)

    kind = "part of the tier ladder (will be replaced by a higher tier)" if exclusive else "permanent (kept even after higher tiers are reached)"
    note = f" Retroactively granted to **{granted}** member(s) who already qualify." if granted else ""
    await log_admin_action(
        interaction.guild, guild_data, interaction.user, "Added XP milestone role",
        target=role.mention, details=f"{xp:,} XP threshold, {'exclusive' if exclusive else 'permanent'}.",
    )
    await interaction.followup.send(
        f"Members reaching **{xp:,} XP** will now automatically receive {role.mention} — {kind}.{note}"
    )


@bot.tree.command(name="removexprole", description="[Admin] Stop auto-assigning a role at an XP milestone")
@app_commands.describe(role="Role to remove from the milestone list")
@app_commands.checks.has_permissions(manage_guild=True)
async def removexprole(interaction: discord.Interaction, role: discord.Role):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    xp_roles = guild_data.setdefault("xp_roles", [])
    before = len(xp_roles)
    xp_roles[:] = [r for r in xp_roles if r["role_id"] != role.id]
    save_data(data)

    if len(xp_roles) < before:
        await log_admin_action(interaction.guild, guild_data, interaction.user, "Removed XP milestone role", target=role.mention)
        await interaction.response.send_message(
            f"{role.mention} will no longer be auto-assigned. Members who already have it keep it — "
            f"this only stops *new* assignments."
        )
    else:
        await interaction.response.send_message("That role wasn't configured as an XP milestone.", ephemeral=True)


@bot.tree.command(name="listxproles", description="Show all configured XP milestone roles")
async def listxproles(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    xp_roles = get_xp_roles(guild_data)

    if not xp_roles:
        await interaction.response.send_message(
            "No XP milestone roles configured yet. Set one up with `/addxprole`.", ephemeral=True
        )
        return

    lines = []
    for mapping in xp_roles:
        role = interaction.guild.get_role(mapping["role_id"])
        role_str = role.mention if role else f"*(deleted role: {mapping['role_id']})*"
        tag = "🔒 sticky" if not mapping["exclusive"] else "🪜 ladder"
        lines.append(f"**{mapping['xp']:,} XP** → {role_str} ({tag})")

    embed = discord.Embed(
        title="🏅 XP Milestone Roles",
        description="\n".join(lines),
        color=discord.Color.teal(),
    )
    embed.set_footer(text="🪜 ladder = only the highest tier is kept  •  🔒 sticky = kept forever once earned")
    await interaction.response.send_message(embed=embed)


@bot.tree.command(
    name="syncxproles",
    description="[Admin] Re-check everyone's roles now, fixing any leftover clutter from before a fix",
)
@app_commands.checks.has_permissions(manage_guild=True)
async def syncxproles(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)

    if not get_xp_roles(guild_data):
        await interaction.response.send_message("No XP milestone roles are configured yet.", ephemeral=True)
        return

    await interaction.response.defer(thinking=True)

    checked = 0
    added_count = 0
    removed_count = 0
    for uid, entry in guild_data["users"].items():
        member = interaction.guild.get_member(int(uid))
        if not member:
            continue
        checked += 1
        result = await apply_xp_roles(interaction.guild, member, guild_data, entry["current_xp"], announce=False)
        added_count += len(result["added"])
        removed_count += len(result["removed"])
    save_data(data)

    await interaction.followup.send(
        f"Re-checked **{checked}** member(s). Added **{added_count}** role(s), "
        f"removed **{removed_count}** redundant lower-tier role(s)."
    )


@bot.tree.command(
    name="setbaseline",
    description="[Admin] Set or correct a user's all-time starting XP (used for the total leaderboard)",
)
@app_commands.describe(
    user="User to adjust",
    starting_xp="Their true starting XP — total gained will be calculated from this point forward",
)
@app_commands.checks.has_permissions(manage_guild=True)
async def setbaseline(interaction: discord.Interaction, user: discord.Member, starting_xp: int):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    entry = guild_data["users"].get(str(user.id))

    if entry is None:
        await interaction.response.send_message(
            "That user has no tracking data yet — have them run `/checkin` first.", ephemeral=True
        )
        return

    entry["baseline_xp"] = starting_xp
    entry["baseline_time"] = iso(utcnow())
    save_data(data)

    total_gained = entry["current_xp"] - starting_xp
    await log_admin_action(
        interaction.guild, guild_data, interaction.user, "Set all-time starting XP",
        target=user.mention, details=f"Baseline set to {starting_xp:,} XP.",
    )
    await interaction.response.send_message(
        f"Set {user.display_name}'s all-time starting XP to **{starting_xp:,}**. "
        f"Total gained is now **{total_gained:,} XP** (from current total of {entry['current_xp']:,})."
    )


@bot.tree.command(name="setweekprogress", description="[Admin] Manually set how far into the week a user is")
@app_commands.describe(
    user="User to adjust",
    days="Days into their week (0-7)",
    hours="Plus additional hours (0-23)",
    minutes="Plus additional minutes (0-59)",
    week_start_xp="Optional: also override their XP total at the start of the week (defaults to current, unchanged)",
)
@app_commands.checks.has_permissions(manage_guild=True)
async def setweekprogress(
    interaction: discord.Interaction,
    user: discord.Member,
    days: int = 0,
    hours: int = 0,
    minutes: int = 0,
    week_start_xp: Optional[int] = None,
):
    delta = timedelta(days=days, hours=hours, minutes=minutes)
    if delta < timedelta(0) or delta > timedelta(days=7):
        await interaction.response.send_message("Total time must be between 0 and 7 days.", ephemeral=True)
        return

    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    entry = guild_data["users"].get(str(user.id))

    if entry is None:
        await interaction.response.send_message("That user has no tracking data.", ephemeral=True)
        return

    new_week_start = utcnow() - delta
    entry["week_start_time"] = iso(new_week_start)
    if week_start_xp is not None:
        entry["week_start_xp"] = week_start_xp
    save_data(data)

    await log_admin_action(
        interaction.guild, guild_data, interaction.user, "Set week progress",
        target=user.mention, details=f"{format_timedelta(delta)} elapsed" + (f", week-start XP set to {week_start_xp:,}" if week_start_xp is not None else ""),
    )
    stats = compute_stats(entry, guild_data["requirement"], get_prorate_threshold_hours(guild_data))
    embed = build_status_embed(user, entry, stats, guild_data["requirement"], show_recent_rate=show_recent_rate_enabled(guild_data))
    embed.description = f"Week progress set to **{format_timedelta(delta)} elapsed** for {user.display_name}."
    await interaction.response.send_message(embed=embed)


@bot.tree.command(
    name="setweekprogressall",
    description="[Admin] Set how far into the week EVERYONE is, in one shot",
)
@app_commands.describe(
    days="Days into the week for all tracked users (0-7)",
    hours="Plus additional hours (0-23)",
    minutes="Plus additional minutes (0-59)",
    reset_gains="If true, zero out everyone's gained-so-far too. Default false = keep existing gains.",
)
@app_commands.checks.has_permissions(manage_guild=True)
async def setweekprogressall(
    interaction: discord.Interaction,
    days: int = 0,
    hours: int = 0,
    minutes: int = 0,
    reset_gains: bool = False,
):
    delta = timedelta(days=days, hours=hours, minutes=minutes)
    if delta < timedelta(0) or delta > timedelta(days=7):
        await interaction.response.send_message("Total time must be between 0 and 7 days.", ephemeral=True)
        return

    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)

    new_week_start = utcnow() - delta
    guild_data["week_anchor"] = iso(new_week_start)

    count = 0
    for entry in guild_data["users"].values():
        entry["week_start_time"] = iso(new_week_start)
        if reset_gains:
            entry["week_start_xp"] = entry["current_xp"]
        count += 1
    save_data(data)

    gains_note = " Everyone's gained-so-far was also reset to 0." if reset_gains else " Existing gained-so-far totals were kept."
    await log_admin_action(
        interaction.guild, guild_data, interaction.user, "Set week progress (everyone)",
        details=f"{format_timedelta(delta)} elapsed for {count} user(s).{gains_note}",
    )
    await interaction.response.send_message(
        f"Set week progress to **{format_timedelta(delta)} elapsed** "
        f"for **{count}** tracked user(s).{gains_note} "
        f"New users who check in for the first time will now also align to this week automatically."
    )


@bot.tree.command(
    name="fullreset",
    description="[Admin] Reset BOTH weekly and all-time totals for a user, starting from current XP",
)
@app_commands.describe(
    user="User to fully reset",
    clear_history="Also wipe their stored checkin history (default: yes)",
)
@app_commands.checks.has_permissions(manage_guild=True)
async def fullreset(interaction: discord.Interaction, user: discord.Member, clear_history: bool = True):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    entry = guild_data["users"].get(str(user.id))

    if entry is None:
        await interaction.response.send_message(
            "That user has no tracking data yet — nothing to reset.", ephemeral=True
        )
        return

    now = utcnow()
    current_xp = entry["current_xp"]

    entry["baseline_xp"] = current_xp
    entry["baseline_time"] = iso(now)
    entry["week_start_time"] = iso(now)
    entry["week_start_xp"] = current_xp
    if clear_history:
        entry["checkins"] = [{"time": iso(now), "xp": current_xp}]

    save_data(data)

    history_note = " Checkin history was also cleared." if clear_history else " Checkin history was kept."
    await log_admin_action(
        interaction.guild, guild_data, interaction.user, "Full reset",
        target=user.mention, details=f"Reset to {current_xp:,} XP.{history_note}",
    )
    await interaction.response.send_message(
        f"Fully reset {user.display_name} — both weekly and all-time totals now start from their "
        f"current XP of **{current_xp:,}**.{history_note}"
    )


@bot.tree.command(name="removeallusers", description="[Admin] Delete tracking data for EVERYONE — asks for confirmation first")
@app_commands.checks.has_permissions(manage_guild=True)
async def removeallusers(interaction: discord.Interaction):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    count = len(guild_data["users"])

    if count == 0:
        await interaction.response.send_message("No one is being tracked yet — nothing to remove.", ephemeral=True)
        return

    view = ConfirmRemoveAllView(interaction.guild_id, count)
    await interaction.response.send_message(
        f"⚠️ This will permanently delete tracking data for **all {count} tracked user(s)** "
        f"in this server — weekly progress, all-time totals, and checkin history. This can't be undone. Continue?",
        view=view,
    )
    view.message = await interaction.original_response()


@bot.tree.command(name="removeuser", description="[Admin] Delete all tracking data for a user")
@app_commands.describe(user="User to remove")
@app_commands.checks.has_permissions(manage_guild=True)
async def removeuser(interaction: discord.Interaction, user: discord.Member):
    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    if str(user.id) in guild_data["users"]:
        del guild_data["users"][str(user.id)]
        save_data(data)
        await log_admin_action(interaction.guild, guild_data, interaction.user, "Removed user", target=user.mention)
        await interaction.response.send_message(f"Removed tracking data for {user.display_name}.")
    else:
        await interaction.response.send_message("No data found for that user.", ephemeral=True)


@bot.tree.command(
    name="removeuserbyid",
    description="[Admin] Remove a tracked user's data by Discord ID — works even if they've left",
)
@app_commands.describe(discord_id="Their numeric Discord user ID (right-click their name → Copy User ID)")
@app_commands.checks.has_permissions(manage_guild=True)
async def removeuserbyid(interaction: discord.Interaction, discord_id: str):
    try:
        uid = str(int(discord_id.strip()))
    except ValueError:
        await interaction.response.send_message("That doesn't look like a valid Discord ID (should be all digits).", ephemeral=True)
        return

    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)
    if uid in guild_data["users"]:
        del guild_data["users"][uid]
        save_data(data)
        member = interaction.guild.get_member(int(uid))
        name = member.display_name if member else f"user {uid} (no longer in this server)"
        await log_admin_action(interaction.guild, guild_data, interaction.user, "Removed user by ID", target=name)
        await interaction.response.send_message(f"Removed tracking data for {name}.")
    else:
        await interaction.response.send_message("No tracking data found for that ID.", ephemeral=True)


@bot.tree.command(
    name="removestaleusers",
    description="[Admin] Bulk-remove tracking data for everyone no longer in this server",
)
@app_commands.checks.has_permissions(manage_guild=True)
async def removestaleusers(interaction: discord.Interaction):
    await interaction.response.defer(thinking=True)

    data = load_data()
    guild_data = get_guild(data, interaction.guild_id)

    if not interaction.guild.chunked:
        try:
            await interaction.guild.chunk()
        except discord.ClientException:
            pass  # members intent not enabled — best-effort with whatever's cached

    stale_uids = [uid for uid in guild_data["users"] if interaction.guild.get_member(int(uid)) is None]

    if not stale_uids:
        await interaction.followup.send("No stale entries found — everyone tracked is still in the server.", ephemeral=True)
        return

    view = ConfirmRemoveStaleView(interaction.guild_id, stale_uids)
    await interaction.followup.send(
        f"⚠️ Found **{len(stale_uids)}** tracked user(s) no longer in this server. "
        f"This will permanently delete their tracking data. Continue?",
        view=view,
    )
    view.message = await interaction.original_response()


# Shared error handler for admin-only commands
async def _perm_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    if isinstance(error, app_commands.errors.MissingPermissions):
        await interaction.response.send_message(
            "You need the **Manage Server** permission to use this command.", ephemeral=True
        )
    else:
        raise error

# Shared error handler for the require_tracking_channel() check
async def _channel_error(interaction: discord.Interaction, error: app_commands.AppCommandError):
    if isinstance(error, app_commands.CheckFailure):
        await interaction.response.send_message(str(error), ephemeral=True)
    else:
        raise error

setrequirement.error(_perm_error)
setchannel.error(_perm_error)
clearchannel.error(_perm_error)
setannouncechannel.error(_perm_error)
clearannouncechannel.error(_perm_error)
setweeklypost.error(_perm_error)
clearweeklypost.error(_perm_error)
setinactivitychannel.error(_perm_error)
clearinactivitychannel.error(_perm_error)
setadminlogchannel.error(_perm_error)
clearadminlogchannel.error(_perm_error)
setinactivitythreshold.error(_perm_error)
toggleinactivitybehindpace.error(_perm_error)
toggleleaderboardpagination.error(_perm_error)
exportdata.error(_perm_error)
undoimport.error(_perm_error)
backup_cmd.error(_perm_error)
restore_cmd.error(_perm_error)
togglerecentrate.error(_perm_error)
togglecompactleaderboard.error(_perm_error)
setproratethreshold.error(_perm_error)
addxprole.error(_perm_error)
removexprole.error(_perm_error)
syncxproles.error(_perm_error)
importxp.error(_perm_error)
settings_cmd.error(_perm_error)
setbaseline.error(_perm_error)
fullreset.error(_perm_error)
resetweek.error(_perm_error)
setweekprogress.error(_perm_error)
setweekprogressall.error(_perm_error)
removeuser.error(_perm_error)
removeuserbyid.error(_perm_error)
removestaleusers.error(_perm_error)

checkin.error(_channel_error)
status.error(_channel_error)
weeklyleaderboard.error(_channel_error)
totalleaderboard.error(_channel_error)
history.error(_channel_error)
progresschart.error(_channel_error)
crewtotals.error(_channel_error)
undo.error(_channel_error)
removeallusers.error(_perm_error)


if __name__ == "__main__":
    token = os.environ.get("DISCORD_TOKEN")
    if not token:
        raise SystemExit(
            "No DISCORD_TOKEN found. Set it as an environment variable or in a .env file."
        )
    bot.run(token)
